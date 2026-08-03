from __future__ import annotations

import re
from collections import defaultdict
from dataclasses import asdict, dataclass, field
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import (
    AuditEvent,
    CaseStatus,
    Contact,
    Customer,
    Product,
    ProductCategory,
    SalesCase,
)
from app.deliverability import validate_address_format
from app.history import reconcile_email_history
from app.product_catalog import (
    category_interest_entries,
    interest_entry,
    merge_customer_interests,
)
from app.products import canonical_product_code

EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,63}", re.IGNORECASE)
SPLIT_PRODUCT_PATTERN = re.compile(r"[,;，、/\n]+")

COMPANY_HEADER = "\u516c\u53f8\u540d\u79f0"
CONTACT_HEADER = "\u8054\u7cfb\u4eba"
EMAIL_HEADER = "\u90ae\u7bb1"
OTHER_EMAIL_HEADER = "\u5176\u4ed6\u90ae\u7bb1"
PRODUCT_HEADER = "\u9700\u8981\u7684Lanya\u4ea7\u54c1(\u724c\u53f7)"
FIRST_CONTACT_HEADER = "\u9996\u6b21\u63a5\u89e6"
LAST_CONTACT_HEADER = "\u6700\u8fd1\u8054\u7cfb"
NO_AI_HEADER = "\u65e0\u9700ai\u53d1\u4fe1"

REQUIRED_HEADERS = (
    COMPANY_HEADER,
    CONTACT_HEADER,
    EMAIL_HEADER,
    OTHER_EMAIL_HEADER,
    FIRST_CONTACT_HEADER,
    LAST_CONTACT_HEADER,
)
GENERIC_CONTACT_NAMES = {"", "customer", "sir/madam", "sir or madam"}
TRUE_VALUES = {"1", "true", "yes", "y", "\u662f", "\u9700\u8981"}


@dataclass(frozen=True)
class SourceAssociation:
    source_row: int
    source_column: str
    company_name: str
    contact_name: str
    product_text: str
    first_contact_at: datetime | None
    last_contact_at: datetime | None
    no_ai: bool
    field_address_count: int


@dataclass
class ParsedEmailEndpoint:
    email: str
    associations: list[SourceAssociation] = field(default_factory=list)

    @property
    def preferred_company_name(self) -> str:
        primary = [
            association.company_name
            for association in self.associations
            if association.source_column == EMAIL_HEADER and association.company_name
        ]
        values = primary or [
            association.company_name
            for association in self.associations
            if association.company_name
        ]
        return values[0] if values else f"Unspecified customer ({self.email})"

    @property
    def preferred_contact_name(self) -> str:
        candidates = {
            association.contact_name.strip()
            for association in self.associations
            if association.source_column == EMAIL_HEADER
            and association.field_address_count == 1
            and association.contact_name.strip()
        }
        return next(iter(candidates)) if len(candidates) == 1 else "Customer"

    @property
    def first_contact_at(self) -> datetime | None:
        values = [
            association.first_contact_at
            for association in self.associations
            if association.first_contact_at is not None
        ]
        return min(values) if values else None

    @property
    def last_contact_at(self) -> datetime | None:
        values = [
            association.last_contact_at
            for association in self.associations
            if association.last_contact_at is not None
        ]
        return max(values) if values else None

    @property
    def suppressed(self) -> bool:
        return any(association.no_ai for association in self.associations)


@dataclass
class ParsedCustomerWorkbook:
    source_file: str
    source_sheet: str
    source_rows: int
    rows_with_email: int
    address_occurrences: int
    endpoints: dict[str, ParsedEmailEndpoint]
    unparsed_email_cells: list[dict[str, Any]]


@dataclass
class FullCustomerImportResult:
    apply: bool
    source_rows: int
    rows_with_email: int
    address_occurrences: int
    unique_addresses: int
    duplicate_address_occurrences: int
    valid_format_addresses: int
    invalid_format_addresses: int
    unparsed_email_cells: int
    existing_addresses: int = 0
    existing_duplicate_addresses: int = 0
    new_addresses: int = 0
    new_companies: int = 0
    created_customers: int = 0
    created_contacts: int = 0
    updated_contacts: int = 0
    created_cases: int = 0
    contact_only_addresses: int = 0
    source_associations: int = 0
    unparsed_samples: list[dict[str, Any]] = field(default_factory=list)

    def to_dict(self) -> dict[str, Any]:
        return asdict(self)


def _as_utc(
    value: Any,
    timezone: ZoneInfo,
    *,
    epoch: datetime = WINDOWS_EPOCH,
) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = from_excel(value, epoch=epoch)
        if isinstance(parsed, time):
            raise ValueError("numeric Excel value does not contain a calendar date")
        if float(value).is_integer():
            parsed = datetime.combine(parsed.date(), time(hour=12))
    elif isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time(hour=12))
    else:
        text = str(value).strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            parsed = datetime.combine(date.fromisoformat(text), time(hour=12))
        else:
            parsed = datetime.fromisoformat(text)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone)
    return parsed.astimezone(UTC)


def _bool(value: Any) -> bool:
    return str(value or "").strip().casefold() in TRUE_VALUES


def _extract_addresses(value: Any) -> tuple[list[str], str | None]:
    text = str(value or "").strip()
    addresses = list(
        dict.fromkeys(match.group(0).strip().casefold() for match in EMAIL_PATTERN.finditer(text))
    )
    residual = EMAIL_PATTERN.sub(" ", text)
    return addresses, residual.strip() if "@" in residual else None


def read_full_customer_workbook(
    workbook_path: Path,
    timezone: ZoneInfo,
) -> ParsedCustomerWorkbook:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    columns = {name: index for index, name in enumerate(headers)}
    missing = sorted(set(REQUIRED_HEADERS) - columns.keys())
    if missing:
        raise ValueError(f"workbook is missing columns: {', '.join(missing)}")

    endpoints: dict[str, ParsedEmailEndpoint] = {}
    unparsed: list[dict[str, Any]] = []
    source_rows = 0
    rows_with_email = 0
    address_occurrences = 0
    for source_row, row in enumerate(rows, start=2):
        source_rows += 1
        first = _as_utc(
            row[columns[FIRST_CONTACT_HEADER]],
            timezone,
            epoch=workbook.epoch,
        )
        last = _as_utc(
            row[columns[LAST_CONTACT_HEADER]],
            timezone,
            epoch=workbook.epoch,
        )
        if first is not None and last is not None and first > last:
            raise ValueError(f"row {source_row}: first contact is later than last contact")
        company_name = str(row[columns[COMPANY_HEADER]] or "").strip()
        contact_name = str(row[columns[CONTACT_HEADER]] or "").strip()
        product_text = (
            str(row[columns[PRODUCT_HEADER]] or "").strip()
            if PRODUCT_HEADER in columns
            else ""
        )
        no_ai = _bool(row[columns[NO_AI_HEADER]]) if NO_AI_HEADER in columns else False
        row_addresses: set[str] = set()
        for header in (EMAIL_HEADER, OTHER_EMAIL_HEADER):
            addresses, residual = _extract_addresses(row[columns[header]])
            if residual:
                unparsed.append(
                    {
                        "row": source_row,
                        "column": header,
                        "value": str(row[columns[header]] or "")[:500],
                    }
                )
            for address in addresses:
                row_addresses.add(address)
                address_occurrences += 1
                endpoint = endpoints.setdefault(address, ParsedEmailEndpoint(email=address))
                association = SourceAssociation(
                    source_row=source_row,
                    source_column=header,
                    company_name=company_name,
                    contact_name=contact_name,
                    product_text=product_text,
                    first_contact_at=first,
                    last_contact_at=last,
                    no_ai=no_ai,
                    field_address_count=len(addresses),
                )
                if association not in endpoint.associations:
                    endpoint.associations.append(association)
        rows_with_email += int(bool(row_addresses))

    return ParsedCustomerWorkbook(
        source_file=workbook_path.name,
        source_sheet=sheet.title,
        source_rows=source_rows,
        rows_with_email=rows_with_email,
        address_occurrences=address_occurrences,
        endpoints=endpoints,
        unparsed_email_cells=unparsed,
    )


def _association_payload(
    association: SourceAssociation,
    parsed: ParsedCustomerWorkbook,
) -> dict[str, Any]:
    return {
        "source_file": parsed.source_file,
        "source_sheet": parsed.source_sheet,
        "source_row": association.source_row,
        "source_column": association.source_column,
        "company_name": association.company_name,
        "contact_name": association.contact_name,
        "product_text": association.product_text,
        "first_contact_at": (
            association.first_contact_at.isoformat()
            if association.first_contact_at is not None
            else None
        ),
        "last_contact_at": (
            association.last_contact_at.isoformat()
            if association.last_contact_at is not None
            else None
        ),
        "no_ai": association.no_ai,
    }


def _merge_contact_metadata(
    contact: Contact,
    endpoint: ParsedEmailEndpoint,
    parsed: ParsedCustomerWorkbook,
) -> None:
    metadata = dict(contact.metadata_json or {})
    existing = list(metadata.get("source_associations") or [])
    known = {
        (
            item.get("source_file"),
            item.get("source_sheet"),
            item.get("source_row"),
            item.get("source_column"),
            item.get("company_name"),
        ): index
        for index, item in enumerate(existing)
        if isinstance(item, dict)
    }
    for association in endpoint.associations:
        payload = _association_payload(association, parsed)
        key = (
            payload["source_file"],
            payload["source_sheet"],
            payload["source_row"],
            payload["source_column"],
            payload["company_name"],
        )
        if key not in known:
            existing.append(payload)
            known[key] = len(existing) - 1
        else:
            existing[known[key]] = payload
    metadata["identity_kind"] = "EMAIL_ENDPOINT"
    metadata.setdefault("identity_verified", False)
    metadata["source_associations"] = existing
    contact.metadata_json = metadata


def _merge_activity(contact: Contact, endpoint: ParsedEmailEndpoint) -> None:
    first_values = [
        value
        for value in (contact.first_contact_at, endpoint.first_contact_at)
        if value is not None
    ]
    last_values = [
        value
        for value in (contact.last_contact_at, endpoint.last_contact_at)
        if value is not None
    ]
    contact.first_contact_at = min(first_values) if first_values else None
    contact.last_contact_at = max(last_values) if last_values else None


def _resolved_products_from_text(
    text: str,
    products_by_code: dict[str, Product],
) -> list[Product]:
    resolved: dict[int, Product] = {}
    stripped = text.strip()
    if not stripped:
        return []
    candidates = [stripped, *SPLIT_PRODUCT_PATTERN.split(stripped)]
    for candidate in candidates:
        code = canonical_product_code(candidate)
        product = products_by_code.get(code.casefold())
        if product is not None:
            resolved[product.id] = product
    return list(resolved.values())


def _resolved_products(
    endpoint: ParsedEmailEndpoint,
    products_by_code: dict[str, Product],
) -> list[Product]:
    resolved: dict[int, Product] = {}
    for association in endpoint.associations:
        for product in _resolved_products_from_text(
            association.product_text,
            products_by_code,
        ):
            resolved[product.id] = product
    return list(resolved.values())


def _generic_name(value: str) -> bool:
    return value.strip().casefold() in GENERIC_CONTACT_NAMES


async def import_full_customer_workbook(
    path: Path,
    session: AsyncSession,
    *,
    apply: bool = False,
    timezone: str = "Asia/Kolkata",
    enable_auto_send: bool = False,
    create_cases: bool = True,
    allow_unparsed_email_cells: bool = False,
    actor: str = "full-customer-import",
) -> FullCustomerImportResult:
    parsed = read_full_customer_workbook(path, ZoneInfo(timezone))
    if apply and parsed.unparsed_email_cells and not allow_unparsed_email_cells:
        raise ValueError(
            "source contains email-like cells that were not fully parsed; "
            "review preview or pass allow_unparsed_email_cells explicitly"
        )

    customers = (
        (await session.execute(select(Customer).order_by(Customer.id))).scalars().all()
    )
    contacts = (
        (await session.execute(select(Contact).order_by(Contact.id))).scalars().all()
    )
    products = (
        (
            await session.execute(
                select(Product).where(Product.active.is_(True)).order_by(Product.id)
            )
        )
        .scalars()
        .all()
    )
    product_categories = (
        (
            await session.execute(
                select(ProductCategory)
                .where(ProductCategory.active.is_(True))
                .order_by(ProductCategory.id)
            )
        )
        .scalars()
        .all()
    )
    categories_by_id = {category.id: category for category in product_categories}
    category_names = {
        category.key: category.name for category in product_categories
    }
    customers_by_name = {
        customer.company_name.strip().casefold(): customer for customer in customers
    }
    contacts_by_email: dict[str, list[Contact]] = defaultdict(list)
    for contact in contacts:
        contacts_by_email[contact.email.strip().casefold()].append(contact)
    products_by_code = {product.code.strip().casefold(): product for product in products}

    valid_formats = sum(
        validate_address_format(address).valid for address in parsed.endpoints
    )
    existing_addresses = sum(
        address in contacts_by_email for address in parsed.endpoints
    )
    existing_duplicate_addresses = sum(
        len(contacts_by_email.get(address, [])) > 1 for address in parsed.endpoints
    )
    new_company_names = {
        endpoint.preferred_company_name.strip().casefold()
        for address, endpoint in parsed.endpoints.items()
        if address not in contacts_by_email
        and endpoint.preferred_company_name.strip().casefold() not in customers_by_name
    }
    result = FullCustomerImportResult(
        apply=apply,
        source_rows=parsed.source_rows,
        rows_with_email=parsed.rows_with_email,
        address_occurrences=parsed.address_occurrences,
        unique_addresses=len(parsed.endpoints),
        duplicate_address_occurrences=parsed.address_occurrences - len(parsed.endpoints),
        valid_format_addresses=valid_formats,
        invalid_format_addresses=len(parsed.endpoints) - valid_formats,
        unparsed_email_cells=len(parsed.unparsed_email_cells),
        existing_addresses=existing_addresses,
        existing_duplicate_addresses=existing_duplicate_addresses,
        new_addresses=len(parsed.endpoints) - existing_addresses,
        new_companies=len(new_company_names),
        source_associations=sum(
            len(endpoint.associations) for endpoint in parsed.endpoints.values()
        ),
        unparsed_samples=parsed.unparsed_email_cells[:20],
    )
    if not apply:
        return result

    case_keys = {
        (row.contact_id, row.product_id, row.currency)
        for row in (
            await session.execute(
                select(
                    SalesCase.contact_id,
                    SalesCase.product_id,
                    SalesCase.currency,
                )
            )
        ).all()
    }
    for address, endpoint in parsed.endpoints.items():
        company_key = endpoint.preferred_company_name.strip().casefold()
        company = customers_by_name.get(company_key)
        matches = contacts_by_email.get(address, [])
        selected_contact = next(
            (
                candidate
                for candidate in matches
                if company is not None and candidate.customer_id == company.id
            ),
            matches[0] if matches else None,
        )
        if selected_contact is None:
            if company is None:
                company = Customer(
                    company_name=endpoint.preferred_company_name,
                    language="en",
                    auto_send_allowed=enable_auto_send,
                    consent_basis="existing CRM/customer-list relationship",
                    do_not_contact=False,
                    metadata_json={
                        "import_source": "full_customer_workbook",
                        "source_file": parsed.source_file,
                    },
                )
                session.add(company)
                await session.flush()
                customers_by_name[company_key] = company
                result.created_customers += 1
            elif enable_auto_send and not company.do_not_contact:
                company.auto_send_allowed = True
                company.consent_basis = (
                    company.consent_basis
                    or "existing CRM/customer-list relationship"
                )
            selected_contact = Contact(
                customer_id=company.id,
                name=endpoint.preferred_contact_name,
                email=address,
                language="en",
                suppressed=endpoint.suppressed,
                metadata_json={},
                first_contact_at=endpoint.first_contact_at,
                last_contact_at=endpoint.last_contact_at,
            )
            session.add(selected_contact)
            await session.flush()
            contacts_by_email[address].append(selected_contact)
            result.created_contacts += 1
        else:
            company = await session.get(Customer, selected_contact.customer_id)
            if (
                company is not None
                and enable_auto_send
                and not company.do_not_contact
            ):
                company.auto_send_allowed = True
                company.consent_basis = (
                    company.consent_basis
                    or "existing CRM/customer-list relationship"
                )
            if _generic_name(selected_contact.name) and not _generic_name(
                endpoint.preferred_contact_name
            ):
                selected_contact.name = endpoint.preferred_contact_name
            selected_contact.suppressed = (
                selected_contact.suppressed or endpoint.suppressed
            )
            _merge_activity(selected_contact, endpoint)
            result.updated_contacts += 1
        _merge_contact_metadata(selected_contact, endpoint, parsed)
        interest_entries: list[dict[str, Any]] = []
        for association in endpoint.associations:
            if not association.product_text.strip():
                continue
            interest_entries.extend(
                category_interest_entries(
                    text=association.product_text,
                    category_names=category_names,
                    source="full_customer_workbook",
                    source_row=association.source_row,
                )
            )
            for product in _resolved_products_from_text(
                association.product_text,
                products_by_code,
            ):
                category = categories_by_id.get(product.category_id)
                if category is None:
                    continue
                interest_entries.append(
                    interest_entry(
                        category_key=category.key,
                        category_name=category.name,
                        source="full_customer_workbook",
                        value=association.product_text,
                        source_row=association.source_row,
                    )
                )
        merge_customer_interests(company, interest_entries)

        resolved_products = (
            _resolved_products(endpoint, products_by_code) if create_cases else []
        )
        if not resolved_products:
            result.contact_only_addresses += 1
        for product in resolved_products:
            key = (selected_contact.id, product.id, "INR")
            if key in case_keys:
                continue
            session.add(
                SalesCase(
                    customer_id=selected_contact.customer_id,
                    contact_id=selected_contact.id,
                    product_id=product.id,
                    category_id=product.category_id,
                    currency="INR",
                    status=CaseStatus.ACTIVE,
                    subject_key=f"{product.name} quotation".casefold(),
                )
            )
            case_keys.add(key)
            result.created_cases += 1

    session.add(
        AuditEvent(
            case_id=None,
            actor=actor,
            event_type="contacts.full_workbook_imported",
            data=result.to_dict(),
        )
    )
    await session.commit()
    await reconcile_email_history(session)
    return result
