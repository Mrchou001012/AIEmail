import csv
import hashlib
import re
from dataclasses import dataclass, field
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

import yaml
from email_validator import EmailNotValidError, validate_email
from openpyxl import Workbook, load_workbook
from sqlalchemy import delete, or_, select, update
from sqlalchemy.ext.asyncio import AsyncSession

from app.commercial import get_or_create_current_cycle, lock_commercial_scope
from app.db import (
    AuditEvent,
    CaseStatus,
    CommercialDataCycle,
    Contact,
    Customer,
    DeliveryStatus,
    InventorySnapshot,
    Outbox,
    PricePolicy,
    Product,
    Quote,
    SalesCase,
)
from app.history import reconcile_email_history
from app.products import canonical_product_code, product_text_key
from app.settings import get_settings

CUSTOMER_HEADERS = [
    "company_name",
    "contact_name",
    "email",
    "language",
    "product_code",
    "currency",
    "auto_send_allowed",
    "consent_basis",
    "do_not_contact",
]
PRICE_HEADERS = [
    "product_code",
    "product_name",
    "approved_text_key",
    "margin_class",
    "currency",
    "unit",
    "standard_price",
    "absolute_floor",
    "max_discount_pct",
    "max_negotiation_rounds",
    "concession_step_pct",
    "min_quantity",
    "max_quantity",
    "tier_1_max_multiple",
    "tier_1_markup_pct",
    "tier_2_max_multiple",
    "tier_2_markup_pct",
    "quote_valid_days",
    "quote_valid_weekday",
    "standard_incoterm",
    "allowed_incoterms",
    "standard_payment_term",
    "allowed_payment_terms",
    "taxes_included",
    "freight_included",
    "manual_only",
    "valid_from",
    "valid_to",
]


@dataclass
class ImportResult:
    source_hash: str
    total_rows: int = 0
    valid_rows: int = 0
    applied_rows: int = 0
    case_ready_rows: int = 0
    contact_only_rows: int = 0
    created_customers: int = 0
    created_contacts: int = 0
    created_cases: int = 0
    missing_product_codes: list[str] = field(default_factory=list)
    errors: list[dict[str, Any]] = field(default_factory=list)
    commercial_cycle_id: int | None = None
    already_applied: bool = False
    inventory_confirmation_required: bool = False

    @property
    def ok(self) -> bool:
        return not self.errors


@dataclass(frozen=True)
class ContentBundle:
    company_profile: str
    product_snippets: dict[str, str]
    compliance_snippets: dict[str, str]
    signature_text: str
    signature_html: str


def load_content(content_dir: Path) -> ContentBundle:
    return ContentBundle(
        company_profile=(content_dir / "company_profile.md").read_text(encoding="utf-8"),
        product_snippets=yaml.safe_load((content_dir / "approved_product_text.yaml").read_text(encoding="utf-8")),
        compliance_snippets=yaml.safe_load((content_dir / "compliance_whitelist.yaml").read_text(encoding="utf-8")),
        signature_text=(content_dir / "email_signature.txt").read_text(encoding="utf-8"),
        signature_html=(content_dir / "email_signature.html").read_text(encoding="utf-8"),
    )


def generate_templates(target_dir: Path) -> None:
    target_dir.mkdir(parents=True, exist_ok=True)
    customer = Workbook()
    ws = customer.active
    ws.title = "customers"
    ws.append(CUSTOMER_HEADERS)
    ws.append(
        [
            "Demo Industrial Ltd",
            "Alex Buyer",
            "buyer@example.com",
            "en",
            "WIDGET-100",
            "USD",
            False,
            "existing business relationship",
            False,
        ]
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = "A1:I2"
    customer.save(target_dir / "customer_list_template.xlsx")

    prices = Workbook()
    ws = prices.active
    ws.title = "prices"
    ws.append(PRICE_HEADERS)
    ws.append(
        [
            "WIDGET-100",
            "Industrial Widget 100",
            "widget_100",
            "A",
            "USD",
            "piece",
            "100.00",
            "82.00",
            "0.15",
            2,
            "0.03",
            10,
            10000,
            4,
            "0.25",
            12,
            "0.20",
            30,
            "FRIDAY",
            "EXW",
            "EXW,FCA,FOB",
            "100% before shipment",
            "100% before shipment,30% deposit / 70% before shipment",
            False,
            False,
            False,
            date.today(),
            None,
        ]
    )
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{ws.cell(1, len(PRICE_HEADERS)).column_letter}2"
    prices.save(target_dir / "price_list_template.xlsx")


def _hash_file(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _rows(path: Path) -> list[dict[str, Any]]:
    if path.suffix.lower() == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            return [dict(row) for row in csv.DictReader(handle) if any(value not in {None, ""} for value in row.values())]
    workbook = load_workbook(path, read_only=True, data_only=True)
    ws = workbook.active
    values = list(ws.iter_rows(values_only=True))
    if not values:
        return []
    headers = [str(value).strip() if value is not None else "" for value in values[0]]
    return [dict(zip(headers, row, strict=False)) for row in values[1:] if any(v is not None for v in row)]


def _bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y"}


def _date(value: Any) -> date | None:
    if value in {None, ""}:
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return date.fromisoformat(str(value))


WEEKDAYS = {
    "MONDAY": 0,
    "TUESDAY": 1,
    "WEDNESDAY": 2,
    "THURSDAY": 3,
    "FRIDAY": 4,
    "SATURDAY": 5,
    "SUNDAY": 6,
}


def _weekday(value: Any) -> int | None:
    if value in {None, ""}:
        return None
    normalized = str(value).strip().upper()
    if normalized in WEEKDAYS:
        return WEEKDAYS[normalized]
    numeric = int(normalized)
    if not 0 <= numeric <= 6:
        raise ValueError
    return numeric


async def import_customers(path: Path, session: AsyncSession, apply: bool = False) -> ImportResult:
    rows = _rows(path)
    result = ImportResult(source_hash=_hash_file(path), total_rows=len(rows))
    parsed: list[dict[str, Any]] = []
    product_cache: dict[str, Product | None] = {}
    missing_product_codes: set[str] = set()
    for number, row in enumerate(rows, start=2):
        errors: list[str] = []
        company = str(row.get("company_name") or "").strip()
        name = str(row.get("contact_name") or "").strip()
        product_code = canonical_product_code(str(row.get("product_code") or ""))
        currency = str(row.get("currency") or "USD").strip().upper()
        if not re.fullmatch(r"[A-Z]{3}", currency):
            errors.append("currency must be a three-letter code")
        try:
            address = validate_email(str(row.get("email") or ""), check_deliverability=False).normalized
        except EmailNotValidError as exc:
            errors.append(f"invalid email: {exc}")
            address = ""
        product = None
        if product_code:
            if product_code not in product_cache:
                product_cache[product_code] = await session.scalar(
                    select(Product).where(Product.code == product_code, Product.active.is_(True))
                )
            product = product_cache[product_code]
            if product is None:
                missing_product_codes.add(product_code)
        if not company:
            errors.append("company_name is required")
        if not name:
            errors.append("contact_name is required")
        if errors:
            result.errors.append({"row": number, "errors": errors})
            continue
        parsed.append(
            {
                **row,
                "email": address,
                "product": product,
                "product_code": product_code,
                "company": company,
                "name": name,
            }
        )
        if product is None:
            result.contact_only_rows += 1
        else:
            result.case_ready_rows += 1
    result.valid_rows = len(parsed)
    result.missing_product_codes = sorted(missing_product_codes)
    if apply and result.ok:
        customer_cache: dict[str, Customer] = {}
        contact_cache: dict[tuple[int, str], Contact] = {}
        case_cache: dict[tuple[int, int, str], SalesCase | None] = {}
        for row in parsed:
            customer = customer_cache.get(row["company"])
            if customer is None:
                customer = await session.scalar(
                    select(Customer).where(Customer.company_name == row["company"])
                )
            row_auto_send = _bool(row.get("auto_send_allowed"))
            row_do_not_contact = _bool(row.get("do_not_contact"))
            row_consent = str(row.get("consent_basis") or "") or None
            if customer is None:
                customer = Customer(
                    company_name=row["company"],
                    auto_send_allowed=row_auto_send,
                    consent_basis=row_consent,
                    do_not_contact=row_do_not_contact,
                )
                session.add(customer)
                await session.flush()
                result.created_customers += 1
            else:
                # Merge repeated company rows conservatively: any suppression wins,
                # and every row must explicitly allow automation.
                customer.auto_send_allowed = customer.auto_send_allowed and row_auto_send
                customer.do_not_contact = customer.do_not_contact or row_do_not_contact
                customer.consent_basis = customer.consent_basis or row_consent
            customer_cache[row["company"]] = customer
            customer.language = str(row.get("language") or "en")
            contact_key = (customer.id, row["email"])
            contact = contact_cache.get(contact_key)
            if contact is None:
                contact = await session.scalar(
                    select(Contact).where(
                        Contact.customer_id == customer.id,
                        Contact.email == row["email"],
                    )
                )
            if contact is None:
                contact = Contact(
                    customer_id=customer.id,
                    name=row["name"],
                    email=row["email"],
                    language=str(row.get("language") or "en"),
                )
                session.add(contact)
                await session.flush()
                result.created_contacts += 1
            else:
                contact.name = row["name"]
                contact.language = str(row.get("language") or "en")
            contact_cache[contact_key] = contact
            if row["product"] is None:
                result.applied_rows += 1
                continue
            currency = str(row.get("currency") or "USD").strip().upper()
            case_key = (contact.id, row["product"].id, currency)
            if case_key not in case_cache:
                case_cache[case_key] = await session.scalar(
                    select(SalesCase).where(
                        SalesCase.customer_id == customer.id,
                        SalesCase.contact_id == contact.id,
                        SalesCase.product_id == row["product"].id,
                        SalesCase.currency == currency,
                        SalesCase.status.in_(
                            [
                                CaseStatus.ACTIVE,
                                CaseStatus.WAITING_HUMAN,
                                CaseStatus.PAUSED,
                                CaseStatus.HUMAN_TAKEOVER,
                            ]
                        ),
                    )
                )
            sales_case = case_cache[case_key]
            if sales_case is None:
                sales_case = SalesCase(
                    customer_id=customer.id,
                    contact_id=contact.id,
                    product_id=row["product"].id,
                    currency=currency,
                    subject_key=f"{row['product'].name} quotation".lower(),
                )
                session.add(sales_case)
                case_cache[case_key] = sales_case
                result.created_cases += 1
            result.applied_rows += 1
        await session.commit()
        await reconcile_email_history(session)
    return result


async def import_prices(
    path: Path,
    session: AsyncSession,
    apply: bool = False,
    replace_active: bool = False,
    actor: str = "system",
) -> ImportResult:
    rows = _rows(path)
    result = ImportResult(source_hash=_hash_file(path), total_rows=len(rows))
    settings = get_settings()
    business_today = datetime.now(UTC).astimezone(ZoneInfo(settings.business_timezone)).date()
    content = load_content(settings.content_dir)
    parsed: list[dict[str, Any]] = []
    for number, row in enumerate(rows, start=2):
        errors: list[str] = []
        code = canonical_product_code(str(row.get("product_code") or ""))
        currency = str(row.get("currency") or "").strip().upper()
        manual_only = _bool(row.get("manual_only"))
        margin_class = str(row.get("margin_class") or "").strip().upper() or None
        approved_text_key = str(row.get("approved_text_key") or product_text_key(code)).strip()
        if not code:
            errors.append("product_code is required")
        if margin_class not in {None, "A", "B"}:
            errors.append("margin_class must be A, B, or blank")
        if not approved_text_key or not str(content.product_snippets.get(approved_text_key) or "").strip():
            errors.append(f"approved product text is missing for key: {approved_text_key or '<empty>'}")
        try:
            standard = Decimal(str(row.get("standard_price") or "0"))
            floor = Decimal(str(row.get("absolute_floor") or "0"))
            max_discount = Decimal(str(row.get("max_discount_pct") or "0"))
            concession = Decimal(str(row.get("concession_step_pct") or "0"))
            tier_1_max = (
                Decimal(str(row.get("tier_1_max_multiple")))
                if row.get("tier_1_max_multiple") not in {None, ""}
                else None
            )
            tier_1_markup = Decimal(str(row.get("tier_1_markup_pct") or "0"))
            tier_2_max = (
                Decimal(str(row.get("tier_2_max_multiple")))
                if row.get("tier_2_max_multiple") not in {None, ""}
                else None
            )
            tier_2_markup = Decimal(str(row.get("tier_2_markup_pct") or "0"))
        except (InvalidOperation, TypeError):
            errors.append("price and percentage fields must be decimals")
            standard = floor = max_discount = concession = Decimal("0")
            tier_1_max = tier_2_max = None
            tier_1_markup = tier_2_markup = Decimal("0")
        try:
            valid_from = _date(row.get("valid_from")) or business_today
            valid_to = _date(row.get("valid_to"))
        except ValueError:
            errors.append("valid_from and valid_to must be ISO dates")
            valid_from = business_today
            valid_to = None
        try:
            max_rounds = int(
                row.get("max_negotiation_rounds")
                if row.get("max_negotiation_rounds") not in {None, ""}
                else 2
            )
            min_quantity = int(row.get("min_quantity") or 1)
            max_quantity = int(row["max_quantity"]) if row.get("max_quantity") else None
            quote_valid_days = int(row.get("quote_valid_days") or 30)
            quote_valid_weekday = _weekday(row.get("quote_valid_weekday"))
            if max_rounds < 0 or min_quantity < 1 or quote_valid_days < 1:
                raise ValueError
            if max_quantity is not None and max_quantity < min_quantity:
                raise ValueError
        except (TypeError, ValueError):
            errors.append("round, quantity, and validity fields must be valid positive integers")
            max_rounds, min_quantity, max_quantity, quote_valid_days = 2, 1, None, 30
            quote_valid_weekday = None
        if not re.fullmatch(r"[A-Z]{3}", currency):
            errors.append("currency must be a three-letter code")
        if not manual_only and (standard <= 0 or floor <= 0 or floor > standard):
            errors.append("prices must be positive and floor cannot exceed standard price")
        if not Decimal("0") <= max_discount < Decimal("1"):
            errors.append("max_discount_pct must be between 0 and 1")
        if valid_to and valid_to < valid_from:
            errors.append("valid_to cannot precede valid_from")
        if tier_1_markup < 0 or tier_2_markup < 0:
            errors.append("tier markups cannot be negative")
        if (tier_1_max is None) != (tier_2_max is None):
            errors.append("tier_1_max_multiple and tier_2_max_multiple must both be set or both be blank")
        if tier_1_max is not None and tier_2_max is not None and not (
            tier_1_max > 1 and tier_2_max > tier_1_max
        ):
            errors.append("quantity tier multiples must satisfy 1 < tier 1 < tier 2")
        product = await session.scalar(select(Product).where(Product.code == code))
        if product is None and not str(row.get("product_name") or "").strip():
            errors.append("new product requires product_name")
        if product is not None and product.approved_text_key != approved_text_key:
            errors.append(
                f"approved_text_key must match existing product key: {product.approved_text_key}"
            )
        if product is not None and not manual_only and not replace_active:
            overlap = await session.scalar(
                select(PricePolicy.id).where(
                    PricePolicy.product_id == product.id,
                    PricePolicy.currency == currency,
                    PricePolicy.active.is_(True),
                    or_(PricePolicy.valid_to.is_(None), PricePolicy.valid_to >= valid_from),
                    True if valid_to is None else PricePolicy.valid_from <= valid_to,
                )
            )
            if overlap:
                errors.append("overlapping active price policy")
        for prior in parsed:
            if prior["code"] == code and (
                prior["product_name_value"] != str(row.get("product_name") or "").strip()
                or prior["unit_value"] != str(row.get("unit") or "unit").strip()
                or prior["approved_text_key"] != approved_text_key
            ):
                errors.append("repeated product rows must use the same name, unit, and approved_text_key")
                break
            prior_end = prior["valid_to_value"] or date.max
            current_end = valid_to or date.max
            if (
                not manual_only
                and not prior["manual_only"]
                and prior["code"] == code
                and prior["currency"] == currency
                and prior["valid_from_value"] <= current_end
                and valid_from <= prior_end
            ):
                errors.append("overlapping price policy within workbook")
                break
        if errors:
            result.errors.append({"row": number, "errors": errors})
            continue
        parsed.append(
            {
                **row,
                "code": code,
                "currency": currency,
                "manual_only": manual_only,
                "margin_class": margin_class,
                "approved_text_key": approved_text_key,
                "product_name_value": str(row.get("product_name") or "").strip(),
                "unit_value": str(row.get("unit") or "unit").strip(),
                "standard": standard,
                "floor": floor,
                "max_discount": max_discount,
                "concession": concession,
                "max_rounds": max_rounds,
                "min_quantity_value": min_quantity,
                "max_quantity_value": max_quantity,
                "tier_1_max": tier_1_max,
                "tier_1_markup": tier_1_markup,
                "tier_2_max": tier_2_max,
                "tier_2_markup": tier_2_markup,
                "quote_valid_days_value": quote_valid_days,
                "quote_valid_weekday": quote_valid_weekday,
                "valid_from_value": valid_from,
                "valid_to_value": valid_to,
                "product": product,
            }
        )
    result.valid_rows = len(parsed)
    if apply and result.ok:
        await lock_commercial_scope(session, settings.commercial_scope)
        cycle = await get_or_create_current_cycle(session, settings)
        cycle = await session.scalar(
            select(CommercialDataCycle)
            .where(CommercialDataCycle.id == cycle.id)
            .with_for_update()
        )
        if cycle is None:
            raise RuntimeError("current commercial-data cycle disappeared")
        result.commercial_cycle_id = cycle.id
        active_rows = [row for row in parsed if not row["manual_only"]]
        if not active_rows:
            result.errors.append(
                {"row": 0, "errors": ["at least one non-manual price row is required to confirm weekly prices"]}
            )
            await session.rollback()
            return result
        already_applied = await session.scalar(
            select(PricePolicy.id).where(
                PricePolicy.commercial_cycle_id == cycle.id,
                PricePolicy.source_hash == result.source_hash,
                PricePolicy.active.is_(True),
            )
        )
        if already_applied is not None:
            result.already_applied = True
            result.applied_rows = len(parsed)
            result.inventory_confirmation_required = cycle.inventory_status != "CONFIRMED"
            await session.commit()
            return result

        in_flight_quote_id = await session.scalar(
            select(Outbox.id)
            .join(Quote, Quote.id == Outbox.quote_id)
            .join(
                CommercialDataCycle,
                CommercialDataCycle.id == Quote.commercial_cycle_id,
            )
            .where(
                CommercialDataCycle.scope == cycle.scope,
                Outbox.message_kind == "AUTO_QUOTE",
                Outbox.status.in_([DeliveryStatus.CLAIMED, DeliveryStatus.UNKNOWN]),
            )
            .limit(1)
        )
        if in_flight_quote_id is not None:
            result.errors.append(
                {
                    "row": 0,
                    "errors": [
                        "price replacement is temporarily blocked while an automatic quote "
                        f"is in flight (outbox {in_flight_quote_id})"
                    ],
                }
            )
            await session.rollback()
            return result

        deactivated_policies = 0
        if replace_active:
            currencies = sorted({row["currency"] for row in active_rows})
            scoped_cycle_ids = select(CommercialDataCycle.id).where(
                CommercialDataCycle.scope == cycle.scope
            )
            scope_predicate = PricePolicy.commercial_cycle_id.in_(scoped_cycle_ids)
            if cycle.scope == "default":
                scope_predicate = or_(
                    scope_predicate,
                    PricePolicy.commercial_cycle_id.is_(None),
                )
            update_result = await session.execute(
                update(PricePolicy)
                .where(
                    PricePolicy.currency.in_(currencies),
                    PricePolicy.active.is_(True),
                    scope_predicate,
                )
                .values(active=False)
            )
            deactivated_policies = int(update_result.rowcount or 0)

        # Any changed price batch invalidates prior stock confirmation for the
        # same week. The operator (or future WMS adapter) must confirm stock
        # against the new priced-product set before automated quotes resume.
        await session.execute(delete(InventorySnapshot).where(InventorySnapshot.cycle_id == cycle.id))
        cycle.inventory_status = "PENDING"
        cycle.inventory_confirmed_at = None
        cycle.inventory_source_system = None
        cycle.inventory_source_ref = None
        cycle.price_status = "PENDING"
        cycle.price_confirmed_at = None
        product_cache = {
            row["code"]: row["product"]
            for row in parsed
            if row["product"] is not None
        }
        for row in parsed:
            product = product_cache.get(row["code"])
            if product is None:
                product = Product(
                    code=row["code"],
                    name=row["product_name_value"],
                    unit=row["unit_value"],
                    approved_text_key=row["approved_text_key"],
                    margin_class=row["margin_class"],
                )
                session.add(product)
                await session.flush()
                product_cache[row["code"]] = product
            else:
                product.margin_class = row["margin_class"]
            if row["manual_only"]:
                result.applied_rows += 1
                continue
            session.add(
                PricePolicy(
                    commercial_cycle_id=cycle.id,
                    product_id=product.id,
                    currency=row["currency"],
                    standard_price=row["standard"],
                    absolute_floor=row["floor"],
                    max_discount_pct=row["max_discount"],
                    max_negotiation_rounds=row["max_rounds"],
                    concession_step_pct=row["concession"],
                    min_quantity=row["min_quantity_value"],
                    max_quantity=row["max_quantity_value"],
                    tier_1_max_multiple=row["tier_1_max"],
                    tier_1_markup_pct=row["tier_1_markup"],
                    tier_2_max_multiple=row["tier_2_max"],
                    tier_2_markup_pct=row["tier_2_markup"],
                    quote_valid_days=row["quote_valid_days_value"],
                    quote_valid_weekday=row["quote_valid_weekday"],
                    standard_incoterm=str(row.get("standard_incoterm") or "EXW"),
                    allowed_incoterms=[v.strip() for v in str(row.get("allowed_incoterms") or "EXW").split(",")],
                    standard_payment_term=str(row.get("standard_payment_term") or "100% before shipment"),
                    allowed_payment_terms=[v.strip() for v in str(row.get("allowed_payment_terms") or "").split(",") if v.strip()],
                    taxes_included=_bool(row.get("taxes_included")),
                    freight_included=_bool(row.get("freight_included")),
                    valid_from=row["valid_from_value"],
                    valid_to=row["valid_to_value"],
                    source_hash=result.source_hash,
                )
            )
            result.applied_rows += 1
        cycle.price_status = "CONFIRMED"
        cycle.price_confirmed_at = datetime.now(UTC)
        cycle.price_source_system = "price_import"
        cycle.price_source_ref = result.source_hash
        cycle.metadata_json = {
            **(cycle.metadata_json or {}),
            "price_rows": len(active_rows),
            "price_replace_active": replace_active,
        }
        session.add(
            AuditEvent(
                actor=actor,
                event_type="commercial.price_replaced",
                data={
                    "cycle_id": cycle.id,
                    "scope": cycle.scope,
                    "source_hash": result.source_hash,
                    "replace_active": replace_active,
                    "deactivated_policies": deactivated_policies,
                    "new_policies": len(active_rows),
                },
            )
        )
        result.inventory_confirmation_required = True
        await session.commit()
    return result
