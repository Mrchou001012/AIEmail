from __future__ import annotations

import csv
import email
import fnmatch
import hashlib
import html
import json
import logging
import math
import multiprocessing
import os
import queue
import re
import tempfile
import zipfile
from collections import Counter
from collections.abc import Iterator
from dataclasses import dataclass
from datetime import UTC, datetime
from enum import StrEnum
from html.parser import HTMLParser
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree

import yaml

logger = logging.getLogger(__name__)

TOKEN_PATTERN = re.compile(r"[A-Za-z0-9][A-Za-z0-9_.+/-]{1,}|[\u3400-\u9fff]")
SPACE_PATTERN = re.compile(r"[ \t\u00a0]+")
BLANK_PATTERN = re.compile(r"\n{3,}")
SYSTEM_NAMES = {".ds_store", "desktop.ini", "thumbs.db"}


def _safe_unicode(value: str) -> str:
    return value.encode("utf-8", errors="backslashreplace").decode("utf-8")


class Classification(StrEnum):
    CUSTOMER_READY = "customer_ready"
    CUSTOMER_CANDIDATE = "customer_candidate"
    INTERNAL = "internal"
    REVIEW_REQUIRED = "review_required"
    EXCLUDED = "excluded"


@dataclass(frozen=True)
class PolicyRule:
    pattern: str
    classification: Classification
    reason: str


@dataclass(frozen=True)
class KnowledgePolicy:
    rules: tuple[PolicyRule, ...]
    excluded_extensions: frozenset[str]
    extractable_extensions: frozenset[str]
    sensitive_markers: tuple[re.Pattern[str], ...]

    @classmethod
    def load(cls, path: Path) -> KnowledgePolicy:
        payload = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
        rules = tuple(
            PolicyRule(
                pattern=str(row["pattern"]).replace("\\", "/"),
                classification=Classification(str(row["classification"])),
                reason=str(row.get("reason") or "policy rule"),
            )
            for row in payload.get("rules", [])
        )
        if not rules:
            raise ValueError("NAS knowledge policy contains no rules")
        return cls(
            rules=rules,
            excluded_extensions=frozenset(str(value).lower() for value in payload.get("excluded_extensions", [])),
            extractable_extensions=frozenset(str(value).lower() for value in payload.get("extractable_extensions", [])),
            sensitive_markers=tuple(re.compile(str(value)) for value in payload.get("sensitive_markers", [])),
        )

    def classify_path(self, relative_path: str) -> tuple[Classification, str]:
        normalized = relative_path.replace("\\", "/").lstrip("/")
        name = PurePosixPath(normalized).name.lower()
        extension = PurePosixPath(normalized).suffix.lower()
        if name in SYSTEM_NAMES or name.startswith("~$"):
            return Classification.EXCLUDED, "system or temporary file"
        if extension in self.excluded_extensions:
            return Classification.EXCLUDED, f"excluded file type {extension or '(none)'}"
        for rule in self.rules:
            if _path_matches(normalized, rule.pattern):
                return rule.classification, rule.reason
        return Classification.REVIEW_REQUIRED, "no explicit policy rule"


def _path_matches(path: str, pattern: str) -> bool:
    if pattern == "**":
        return True
    if pattern.endswith("/**"):
        prefix = pattern[:-3].rstrip("/")
        return path == prefix or path.startswith(f"{prefix}/")
    return fnmatch.fnmatchcase(path, pattern)


class _TextHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.parts: list[str] = []

    def handle_data(self, data: str) -> None:
        if data.strip():
            self.parts.append(data)


def _normalize_text(value: str) -> str:
    value = html.unescape(value).replace("\r\n", "\n").replace("\r", "\n").replace("\x00", "")
    lines = [SPACE_PATTERN.sub(" ", line).strip() for line in value.splitlines()]
    return _safe_unicode(BLANK_PATTERN.sub("\n\n", "\n".join(lines)).strip())


def _decode_text(raw: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-16", "gb18030", "latin-1"):
        try:
            return raw.decode(encoding)
        except UnicodeDecodeError:
            continue
    return raw.decode("utf-8", errors="replace")


def _xml_text(raw: bytes) -> str:
    root = ElementTree.fromstring(raw)
    return "\n".join(value.strip() for value in root.itertext() if value.strip())


def _extract_office_zip(path: Path, prefix: str, suffix: str) -> str:
    parts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        names = sorted(name for name in archive.namelist() if name.startswith(prefix) and name.endswith(suffix))
        for name in names:
            parts.append(_xml_text(archive.read(name)))
    return _normalize_text("\n\n".join(parts))


def _extract_xlsx(path: Path) -> str:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    parts: list[str] = []
    remaining_cells = 25_000
    try:
        for sheet in workbook.worksheets:
            parts.append(f"[Sheet: {sheet.title}]")
            max_row = min(sheet.max_row or 1, 5_000)
            max_column = min(sheet.max_column or 1, 100)
            for row in sheet.iter_rows(max_row=max_row, max_col=max_column, values_only=True):
                values = [str(value).strip() for value in row if value is not None and str(value).strip()]
                if values:
                    parts.append("\t".join(values))
                remaining_cells -= len(row)
                if remaining_cells <= 0:
                    parts.append("[Extraction truncated at safe workbook cell limit]")
                    return _normalize_text("\n".join(parts))
    finally:
        workbook.close()
    return _normalize_text("\n".join(parts))


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader
    except ImportError as exc:  # pragma: no cover - deployment dependency guard
        raise RuntimeError("pypdf is not installed") from exc
    reader = PdfReader(path)
    pages = reader.pages[:50]
    text = "\n\n".join(page.extract_text() or "" for page in pages)
    if len(reader.pages) > len(pages):
        text += "\n\n[Extraction truncated at safe PDF page limit]"
    return _normalize_text(text)


def extract_document(path: Path) -> str:
    extension = path.suffix.lower()
    if extension in {".txt", ".md", ".json", ".yaml", ".yml", ".csv", ".tsv"}:
        return _normalize_text(_decode_text(path.read_bytes()))
    if extension in {".htm", ".html"}:
        parser = _TextHTMLParser()
        parser.feed(_decode_text(path.read_bytes()))
        return _normalize_text("\n".join(parser.parts))
    if extension == ".eml":
        message = email.message_from_bytes(path.read_bytes())
        parts = [str(message.get("Subject") or "")]
        for part in message.walk():
            if part.get_content_type() != "text/plain" or part.get_content_disposition() == "attachment":
                continue
            payload = part.get_payload(decode=True)
            if payload:
                parts.append(_decode_text(payload))
        return _normalize_text("\n\n".join(parts))
    if extension == ".docx":
        return _extract_office_zip(path, "word/", "document.xml")
    if extension == ".pptx":
        return _extract_office_zip(path, "ppt/slides/slide", ".xml")
    if extension == ".xlsx":
        return _extract_xlsx(path)
    if extension == ".pdf":
        return _extract_pdf(path)
    raise ValueError(f"unsupported document type: {extension or '(none)'}")


def _extract_in_process(path: str, result_queue: Any) -> None:
    try:
        result_queue.put((True, extract_document(Path(path))))
    except BaseException as exc:
        result_queue.put((False, f"{type(exc).__name__}: {str(exc)[:500]}"))


def extract_document_bounded(path: Path, *, timeout_seconds: int) -> str:
    if path.suffix.lower() not in {".pdf", ".xlsx"}:
        return extract_document(path)
    context = multiprocessing.get_context("spawn")
    result_queue = context.Queue(maxsize=1)
    process = context.Process(target=_extract_in_process, args=(str(path), result_queue), daemon=True)
    process.start()
    try:
        succeeded, payload = result_queue.get(timeout=timeout_seconds)
    except queue.Empty as exc:
        process.terminate()
        process.join(timeout=5)
        raise TimeoutError(f"document extraction exceeded {timeout_seconds} seconds") from exc
    finally:
        result_queue.close()
    process.join(timeout=5)
    if process.is_alive():
        process.terminate()
        process.join(timeout=5)
    if not succeeded:
        raise RuntimeError(str(payload))
    return str(payload)


def _chunks(text: str, *, size: int = 1600, overlap: int = 200) -> Iterator[str]:
    if size <= overlap:
        raise ValueError("chunk size must be greater than overlap")
    cursor = 0
    while cursor < len(text):
        end = min(len(text), cursor + size)
        if end < len(text):
            natural = max(text.rfind("\n", cursor + size // 2, end), text.rfind(". ", cursor + size // 2, end))
            if natural > cursor:
                end = natural + 1
        chunk = text[cursor:end].strip()
        if chunk:
            yield chunk
        if end >= len(text):
            break
        cursor = max(cursor + 1, end - overlap)


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        while block := stream.read(1024 * 1024):
            digest.update(block)
    return digest.hexdigest()


def _atomic_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", encoding="utf-8", dir=path.parent, delete=False, suffix=".part") as stream:
        json.dump(payload, stream, ensure_ascii=True, separators=(",", ":"))
        temporary = Path(stream.name)
    temporary.replace(path)


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    return json.loads(path.read_text(encoding="utf-8"))


@dataclass(frozen=True)
class ScanPaths:
    manifest: Path
    index: Path
    overrides: Path
    inventory_csv: Path
    summary: Path

    @classmethod
    def in_directory(cls, directory: Path) -> ScanPaths:
        return cls(
            manifest=directory / "manifest.json",
            index=directory / "knowledge_index.json",
            overrides=directory / "classification_overrides.json",
            inventory_csv=directory / "inventory.csv",
            summary=directory / "classification_summary.json",
        )


class NASKnowledgeScanner:
    def __init__(
        self,
        *,
        root: Path,
        policy_path: Path,
        output_dir: Path,
        max_extract_bytes: int = 50 * 1024 * 1024,
        extraction_timeout_seconds: int = 5,
    ) -> None:
        self.root = root
        root_text = str(root)
        if os.name == "nt" and root_text.startswith("\\\\") and not root_text.startswith("\\\\?\\"):
            self.filesystem_root = Path(f"\\\\?\\UNC\\{root_text[2:]}")
        else:
            self.filesystem_root = root
        self.policy = KnowledgePolicy.load(policy_path)
        self.paths = ScanPaths.in_directory(output_dir)
        self.max_extract_bytes = max_extract_bytes
        self.extraction_timeout_seconds = extraction_timeout_seconds
        self.scan_warnings: list[dict[str, str]] = []

    def _files(self) -> Iterator[tuple[Path, str, os.stat_result]]:
        stack = [self.filesystem_root]
        while stack:
            directory = stack.pop()
            try:
                with os.scandir(directory) as entries:
                    ordered = sorted(entries, key=lambda item: item.name.lower(), reverse=True)
            except OSError as exc:
                logger.warning("cannot enumerate NAS directory %s: %s", directory, exc)
                try:
                    relative = directory.relative_to(self.filesystem_root).as_posix()
                except ValueError:
                    relative = str(directory)
                self.scan_warnings.append({"path": relative, "error": f"{type(exc).__name__}: {str(exc)[:300]}"})
                continue
            for entry in ordered:
                try:
                    if entry.is_symlink():
                        continue
                    if entry.is_dir(follow_symlinks=False):
                        relative_dir = _safe_unicode(Path(entry.path).relative_to(self.filesystem_root).as_posix())
                        classification, _ = self.policy.classify_path(relative_dir)
                        if classification != Classification.EXCLUDED:
                            stack.append(Path(entry.path))
                        continue
                    if not entry.is_file(follow_symlinks=False):
                        continue
                    path = Path(entry.path)
                    relative = _safe_unicode(path.relative_to(self.filesystem_root).as_posix())
                    yield path, relative, entry.stat(follow_symlinks=False)
                except OSError as exc:
                    logger.warning("cannot stat NAS entry %s: %s", entry.path, exc)

    def scan(self) -> dict[str, Any]:
        started = datetime.now(UTC)
        self.scan_warnings = []
        previous_payload = _load_json(self.paths.manifest, {"documents": []})
        previous = {str(row.get("path")): row for row in previous_payload.get("documents", [])}
        previous_index = _load_json(self.paths.index, {"chunks": []})
        previous_chunks: dict[str, list[dict[str, Any]]] = {}
        for row in previous_index.get("chunks", []):
            previous_chunks.setdefault(str(row.get("path")), []).append(row)
        overrides = _load_json(self.paths.overrides, {})
        documents: list[dict[str, Any]] = []
        chunks: list[dict[str, Any]] = []
        errors = 0
        changed = 0
        visited = 0
        current_top_level: str | None = None
        seen_paths: set[str] = set()

        def persist_checkpoint() -> None:
            unseen_paths = previous.keys() - seen_paths
            checkpoint_documents = [*documents, *(previous[path] for path in unseen_paths)]
            checkpoint_chunks = [*chunks]
            for unseen_path in unseen_paths:
                checkpoint_chunks.extend(previous_chunks.get(unseen_path, []))
            self._persist_snapshot(
                documents=checkpoint_documents,
                chunks=checkpoint_chunks,
                started=started,
                changed=changed,
                errors=errors,
                complete=False,
            )

        for path, relative, stat in self._files():
            visited += 1
            next_top_level = relative.split("/", 1)[0]
            if current_top_level is not None and next_top_level != current_top_level:
                persist_checkpoint()
            current_top_level = next_top_level
            if visited == 1 or visited % 250 == 0:
                logger.info("NAS knowledge scan visited %s files; current=%s", visited, relative)
            extension = path.suffix.lower()
            base_classification, reason = self.policy.classify_path(relative)
            override = overrides.get(relative)
            if isinstance(override, dict):
                base_classification = Classification(str(override["classification"]))
                reason = f"manual override: {override.get('reason') or 'administrator decision'}"
            fingerprint = f"{stat.st_size}:{stat.st_mtime_ns}"
            old = previous.get(relative)
            unchanged = bool(
                old
                and old.get("fingerprint") == fingerprint
                and old.get("policy_source") == reason
                and (not int(old.get("chunk_count") or 0) or relative in previous_chunks)
            )
            document: dict[str, Any] = {
                "path": relative,
                "name": _safe_unicode(path.name),
                "extension": extension,
                "size": stat.st_size,
                "modified_at": datetime.fromtimestamp(stat.st_mtime, UTC).isoformat(),
                "fingerprint": fingerprint,
                "classification": base_classification.value,
                "classification_source": reason,
                "policy_source": reason,
                "extract_status": "not_extractable",
                "sha256": None,
                "sensitive_hits": [],
                "chunk_count": 0,
                "error": None,
            }
            old_chunks: list[dict[str, Any]] = []
            if unchanged and old:
                document.update(
                    {
                        key: old.get(key)
                        for key in ("classification", "extract_status", "sha256", "sensitive_hits", "chunk_count", "error")
                    }
                )
                old_chunks = previous_chunks.get(relative, [])
            elif base_classification != Classification.EXCLUDED and extension in self.policy.extractable_extensions:
                changed += 1
                if stat.st_size > self.max_extract_bytes:
                    document["extract_status"] = "too_large"
                    if base_classification == Classification.CUSTOMER_CANDIDATE:
                        document["classification"] = Classification.REVIEW_REQUIRED.value
                        document["classification_source"] = "candidate exceeds safe extraction limit"
                else:
                    try:
                        text = extract_document_bounded(
                            path,
                            timeout_seconds=(
                                max(self.extraction_timeout_seconds, 15)
                                if base_classification in {
                                    Classification.CUSTOMER_CANDIDATE,
                                    Classification.CUSTOMER_READY,
                                }
                                else self.extraction_timeout_seconds
                            ),
                        )
                        document["sha256"] = _sha256(path)
                        if not text:
                            document["extract_status"] = "empty"
                            if base_classification == Classification.CUSTOMER_CANDIDATE:
                                document["classification"] = Classification.REVIEW_REQUIRED.value
                                document["classification_source"] = "candidate contains no extractable text"
                        else:
                            hits = [pattern.pattern for pattern in self.policy.sensitive_markers if pattern.search(text)]
                            document["sensitive_hits"] = hits
                            document["extract_status"] = "indexed"
                            if base_classification == Classification.CUSTOMER_CANDIDATE:
                                if hits:
                                    document["classification"] = Classification.REVIEW_REQUIRED.value
                                    document["classification_source"] = "candidate matched sensitive-content guard"
                                else:
                                    document["classification"] = Classification.CUSTOMER_READY.value
                                    document["classification_source"] = "approved path and local content guard passed"
                            for number, chunk in enumerate(_chunks(text), start=1):
                                old_chunks.append(
                                    {
                                        "id": hashlib.sha256(f"{relative}:{number}:{document['sha256']}".encode()).hexdigest()[:24],
                                        "path": relative,
                                        "chunk": number,
                                        "classification": document["classification"],
                                        "text": chunk,
                                        "tokens": _tokenize(chunk),
                                    }
                                )
                            document["chunk_count"] = len(old_chunks)
                    except Exception as exc:  # a bad source document must not abort the NAS sweep
                        errors += 1
                        document["extract_status"] = "error"
                        document["error"] = f"{type(exc).__name__}: {str(exc)[:300]}"
                        if base_classification == Classification.CUSTOMER_CANDIDATE:
                            document["classification"] = Classification.REVIEW_REQUIRED.value
                            document["classification_source"] = "candidate extraction failed"
            documents.append(document)
            chunks.extend(old_chunks)
            seen_paths.add(relative)
            if visited % 500 == 0:
                persist_checkpoint()

        return self._persist_snapshot(
            documents=documents,
            chunks=chunks,
            started=started,
            changed=changed,
            errors=errors,
            complete=True,
        )

    def _persist_snapshot(
        self,
        *,
        documents: list[dict[str, Any]],
        chunks: list[dict[str, Any]],
        started: datetime,
        changed: int,
        errors: int,
        complete: bool,
    ) -> dict[str, Any]:
        sorted_documents = sorted(documents, key=lambda row: str(row["path"]).lower())
        sorted_chunks = sorted(chunks, key=lambda row: (str(row["path"]).lower(), int(row["chunk"])))
        completed = datetime.now(UTC)
        counts = Counter(str(row["classification"]) for row in sorted_documents)
        extract_counts = Counter(str(row["extract_status"]) for row in sorted_documents)
        summary = {
            "schema_version": "nas-knowledge-summary.v1",
            "root": str(self.root),
            "started_at": started.isoformat(),
            "completed_at": completed.isoformat(),
            "complete": complete,
            "duration_seconds": round((completed - started).total_seconds(), 3),
            "file_count": len(sorted_documents),
            "chunk_count": len(sorted_chunks),
            "changed_file_count": changed,
            "error_count": errors,
            "enumeration_warning_count": len(self.scan_warnings),
            "enumeration_warnings": self.scan_warnings[:200],
            "classification_counts": dict(sorted(counts.items())),
            "extraction_counts": dict(sorted(extract_counts.items())),
        }
        _atomic_json(
            self.paths.manifest,
            {**summary, "schema_version": "nas-knowledge-manifest.v1", "documents": sorted_documents},
        )
        _atomic_json(
            self.paths.index,
            {
                "schema_version": "nas-knowledge-index.v1",
                "created_at": completed.isoformat(),
                "root": str(self.root),
                "complete": complete,
                "chunks": sorted_chunks,
            },
        )
        _atomic_json(self.paths.summary, summary)
        self._write_inventory(sorted_documents)
        return summary

    def _write_inventory(self, documents: list[dict[str, Any]]) -> None:
        self.paths.inventory_csv.parent.mkdir(parents=True, exist_ok=True)
        temporary = self.paths.inventory_csv.with_suffix(".csv.part")
        columns = [
            "path",
            "classification",
            "classification_source",
            "extension",
            "size",
            "modified_at",
            "extract_status",
            "chunk_count",
            "error",
        ]
        with temporary.open("w", encoding="utf-8-sig", errors="backslashreplace", newline="") as stream:
            writer = csv.DictWriter(stream, fieldnames=columns, extrasaction="ignore")
            writer.writeheader()
            writer.writerows(documents)
        temporary.replace(self.paths.inventory_csv)


def _tokenize(text: str) -> list[str]:
    return [token.lower() for token in TOKEN_PATTERN.findall(text)][:2000]


@dataclass(frozen=True)
class KnowledgeMatch:
    path: str
    chunk: int
    score: float
    text: str
    classification: Classification


class LocalKnowledgeBase:
    def __init__(self, index_path: Path) -> None:
        payload = _load_json(index_path, {})
        if payload.get("schema_version") != "nas-knowledge-index.v1":
            raise ValueError("unsupported NAS knowledge index schema")
        self.rows = list(payload.get("chunks") or [])

    def search(
        self,
        query: str,
        *,
        audience: str = "customer",
        top_k: int = 5,
    ) -> tuple[KnowledgeMatch, ...]:
        if audience not in {"customer", "internal"}:
            raise ValueError("audience must be customer or internal")
        if not 1 <= top_k <= 20:
            raise ValueError("top_k must be between 1 and 20")
        query_tokens = _tokenize(query)
        if not query_tokens:
            return ()
        allowed = {Classification.CUSTOMER_READY.value}
        if audience == "internal":
            allowed.add(Classification.INTERNAL.value)
        eligible = [row for row in self.rows if row.get("classification") in allowed]
        if not eligible:
            return ()
        document_frequency = Counter()
        for row in eligible:
            document_frequency.update(set(row.get("tokens") or []))
        scores: list[KnowledgeMatch] = []
        for row in eligible:
            tokens = list(row.get("tokens") or [])
            frequencies = Counter(tokens)
            score = 0.0
            for token in query_tokens:
                if not frequencies[token]:
                    continue
                inverse = math.log((len(eligible) + 1) / (document_frequency[token] + 0.5)) + 1
                score += inverse * (1 + math.log(frequencies[token]))
            text = str(row.get("text") or "")
            normalized_query = query.strip().lower()
            if normalized_query and normalized_query in text.lower():
                score += 4.0
            if score <= 0:
                continue
            scores.append(
                KnowledgeMatch(
                    path=str(row["path"]),
                    chunk=int(row["chunk"]),
                    score=score,
                    text=text,
                    classification=Classification(str(row["classification"])),
                )
            )
        scores.sort(key=lambda item: (item.score, item.path, -item.chunk), reverse=True)
        return tuple(scores[:top_k])


def set_classification_override(
    *,
    output_dir: Path,
    relative_path: str,
    classification: Classification,
    reason: str,
    actor: str,
) -> dict[str, Any]:
    paths = ScanPaths.in_directory(output_dir)
    normalized = relative_path.replace("\\", "/").lstrip("/")
    if ".." in PurePosixPath(normalized).parts or not normalized:
        raise ValueError("invalid relative NAS path")
    overrides = _load_json(paths.overrides, {})
    overrides[normalized] = {
        "classification": classification.value,
        "reason": reason.strip()[:500],
        "actor": actor,
        "updated_at": datetime.now(UTC).isoformat(),
    }
    _atomic_json(paths.overrides, overrides)
    return overrides[normalized]


def read_scan_summary(output_dir: Path) -> dict[str, Any]:
    return _load_json(ScanPaths.in_directory(output_dir).summary, {})


def list_documents(
    output_dir: Path,
    *,
    classification: Classification | None = None,
    limit: int = 200,
) -> list[dict[str, Any]]:
    payload = _load_json(ScanPaths.in_directory(output_dir).manifest, {"documents": []})
    rows = list(payload.get("documents") or [])
    if classification is not None:
        rows = [row for row in rows if row.get("classification") == classification.value]
    return rows[:limit]
