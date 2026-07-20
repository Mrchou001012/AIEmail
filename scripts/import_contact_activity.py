from __future__ import annotations

import argparse
import asyncio
import os
import re
from collections import defaultdict
from datetime import UTC, date, datetime, time
from pathlib import Path
from typing import Any
from zoneinfo import ZoneInfo

from dotenv import dotenv_values
from openpyxl import load_workbook
from openpyxl.utils.datetime import WINDOWS_EPOCH, from_excel

EMAIL_PATTERN = re.compile(r"[A-Z0-9._%+\-]+@[A-Z0-9.\-]+\.[A-Z]{2,}", re.IGNORECASE)
EMAIL_HEADER = "\u90ae\u7bb1"
OTHER_EMAIL_HEADER = "\u5176\u4ed6\u90ae\u7bb1"
FIRST_CONTACT_HEADER = "\u9996\u6b21\u63a5\u89e6"
LAST_CONTACT_HEADER = "\u6700\u8fd1\u8054\u7cfb"
ACTIVITY_HEADERS = (
    EMAIL_HEADER,
    OTHER_EMAIL_HEADER,
    FIRST_CONTACT_HEADER,
    LAST_CONTACT_HEADER,
)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Backfill Contact activity dates from the original Chinese CRM workbook."
    )
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--env-file", type=Path, default=Path("/etc/aiemail/aiemail.env"))
    parser.add_argument("--timezone", default="Asia/Kolkata")
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Persist updates. Without this flag the command is a read-only preview.",
    )
    args = parser.parse_args()
    if not args.workbook.is_file():
        parser.error(f"workbook does not exist: {args.workbook}")
    if not args.env_file.is_file():
        parser.error(f"environment file does not exist: {args.env_file}")
    ZoneInfo(args.timezone)
    return args


def _as_utc(
    value: Any,
    timezone: ZoneInfo,
    *,
    epoch: datetime = WINDOWS_EPOCH,
) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        parsed = from_excel(value, epoch=epoch)
        if isinstance(parsed, time):
            raise ValueError("numeric Excel value does not contain a calendar date")
        if float(value).is_integer():
            parsed = datetime.combine(parsed.date(), time(hour=12))
    elif isinstance(value, datetime):
        parsed = value
    elif isinstance(value, date):
        parsed = datetime.combine(value, time(hour=12))
    else:
        text = str(value).strip()
        if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
            parsed = datetime.combine(date.fromisoformat(text), time(hour=12))
        else:
            parsed = datetime.fromisoformat(text)
    if parsed.tzinfo is None:
        parsed = parsed.replace(tzinfo=timezone)
    return parsed.astimezone(UTC)


def _merge_date(
    target: dict[str, tuple[datetime | None, datetime | None]],
    email: str,
    first: datetime | None,
    last: datetime | None,
) -> None:
    previous_first, previous_last = target.get(email, (None, None))
    first_values = [item for item in (previous_first, first) if item is not None]
    last_values = [item for item in (previous_last, last) if item is not None]
    target[email] = (
        min(first_values) if first_values else None,
        max(last_values) if last_values else None,
    )


def read_activity_dates(
    workbook_path: Path,
    timezone: ZoneInfo,
) -> tuple[dict[str, tuple[datetime | None, datetime | None]], int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = [str(value).strip() if value is not None else "" for value in next(rows)]
    columns = {name: index for index, name in enumerate(headers)}
    missing = sorted(set(ACTIVITY_HEADERS) - columns.keys())
    if missing:
        raise ValueError(f"workbook is missing columns: {', '.join(missing)}")
    activity: dict[str, tuple[datetime | None, datetime | None]] = {}
    row_count = 0
    for source_row, row in enumerate(rows, start=2):
        row_count += 1
        first = _as_utc(
            row[columns[FIRST_CONTACT_HEADER]],
            timezone,
            epoch=workbook.epoch,
        )
        last = _as_utc(
            row[columns[LAST_CONTACT_HEADER]],
            timezone,
            epoch=workbook.epoch,
        )
        if first is not None and last is not None and first > last:
            raise ValueError(f"row {source_row}: first contact is later than last contact")
        email_text = " ".join(
            str(row[columns[name]] or "") for name in (EMAIL_HEADER, OTHER_EMAIL_HEADER)
        )
        for match in EMAIL_PATTERN.findall(email_text):
            _merge_date(activity, match.strip().casefold(), first, last)
    if any(
        first is not None and last is not None and first > last
        for first, last in activity.values()
    ):
        raise ValueError("merged activity contains a first contact later than its last contact")
    return activity, row_count


async def run(args: argparse.Namespace) -> dict[str, int | bool]:
    os.environ.update(
        {
            key: value
            for key, value in dotenv_values(args.env_file).items()
            if value is not None
        }
    )
    from sqlalchemy import select

    from app.db import Contact, SessionLocal

    activity, source_rows = read_activity_dates(args.workbook, ZoneInfo(args.timezone))
    async with SessionLocal() as session:
        contacts = (await session.execute(select(Contact))).scalars().all()
        contacts_by_email: dict[str, list[Contact]] = defaultdict(list)
        for contact in contacts:
            contacts_by_email[contact.email.strip().casefold()].append(contact)
        matched_contacts = 0
        changed_contacts = 0
        matched_addresses = 0
        for email, (first, last) in activity.items():
            matches = contacts_by_email.get(email, [])
            if not matches:
                continue
            matched_addresses += 1
            for contact in matches:
                matched_contacts += 1
                before = (contact.first_contact_at, contact.last_contact_at)
                first_values = [item for item in (contact.first_contact_at, first) if item is not None]
                last_values = [item for item in (contact.last_contact_at, last) if item is not None]
                contact.first_contact_at = min(first_values) if first_values else None
                contact.last_contact_at = max(last_values) if last_values else None
                changed_contacts += int(before != (contact.first_contact_at, contact.last_contact_at))
        if args.apply:
            await session.commit()
        else:
            await session.rollback()
    return {
        "apply": args.apply,
        "source_rows": source_rows,
        "source_addresses": len(activity),
        "matched_addresses": matched_addresses,
        "matched_contacts": matched_contacts,
        "changed_contacts": changed_contacts,
        "unmatched_addresses": len(activity) - matched_addresses,
    }


def main() -> None:
    args = parse_args()
    result = asyncio.run(run(args))
    for key, value in result.items():
        print(f"{key}={value}")


if __name__ == "__main__":
    main()
