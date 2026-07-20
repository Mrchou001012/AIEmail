from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

import pytest
from openpyxl import Workbook
from openpyxl.utils.datetime import MAC_EPOCH, to_excel

from scripts.import_contact_activity import ACTIVITY_HEADERS, read_activity_dates


def test_original_crm_dates_and_secondary_emails_are_parsed(tmp_path: Path) -> None:
    path = tmp_path / "crm.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(ACTIVITY_HEADERS)
    sheet.append(
        [
            "buyer@example.com",
            "second@example.com; third@example.com",
            44334,
            "2025-06-07",
        ]
    )
    workbook.save(path)

    activity, row_count = read_activity_dates(path, ZoneInfo("Asia/Kolkata"))

    assert row_count == 1
    assert set(activity) == {
        "buyer@example.com",
        "second@example.com",
        "third@example.com",
    }
    first, last = activity["buyer@example.com"]
    assert first is not None and first.date() == date(2021, 5, 18)
    assert last is not None and last.date() == date(2025, 6, 7)


def test_original_crm_numeric_dates_use_the_workbook_epoch(tmp_path: Path) -> None:
    path = tmp_path / "crm-1904.xlsx"
    workbook = Workbook()
    workbook.epoch = MAC_EPOCH
    sheet = workbook.active
    sheet.append(ACTIVITY_HEADERS)
    sheet.append(
        [
            "buyer@example.com",
            None,
            to_excel(datetime(2021, 5, 18), epoch=MAC_EPOCH),
            to_excel(datetime(2025, 6, 7), epoch=MAC_EPOCH),
        ]
    )
    workbook.save(path)

    activity, _ = read_activity_dates(path, ZoneInfo("Asia/Kolkata"))

    first, last = activity["buyer@example.com"]
    assert first is not None and first.date() == date(2021, 5, 18)
    assert last is not None and last.date() == date(2025, 6, 7)


def test_original_crm_rejects_invalid_dates_after_duplicate_email_merge(
    tmp_path: Path,
) -> None:
    path = tmp_path / "crm-invalid-merge.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(ACTIVITY_HEADERS)
    sheet.append(["buyer@example.com", None, "2026-01-01", None])
    sheet.append(["buyer@example.com", None, None, "2025-01-01"])
    workbook.save(path)

    with pytest.raises(ValueError, match="merged activity"):
        read_activity_dates(path, ZoneInfo("Asia/Kolkata"))
