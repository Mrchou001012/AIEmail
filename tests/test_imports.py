from pathlib import Path
from unittest.mock import AsyncMock

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.db import Contact, Customer, SalesCase
from app.imports import (
    CUSTOMER_HEADERS,
    PRICE_HEADERS,
    generate_templates,
    import_customers,
    import_prices,
)


def test_template_generation(tmp_path: Path) -> None:
    generate_templates(tmp_path)
    customer = load_workbook(tmp_path / "customer_list_template.xlsx", read_only=True)
    prices = load_workbook(tmp_path / "price_list_template.xlsx", read_only=True)
    assert [cell.value for cell in next(customer.active.iter_rows())] == CUSTOMER_HEADERS
    assert [cell.value for cell in next(prices.active.iter_rows())] == PRICE_HEADERS


@pytest.mark.asyncio
async def test_invalid_customer_email_is_row_error(tmp_path: Path) -> None:
    generate_templates(tmp_path)
    path = tmp_path / "customer_list_template.xlsx"
    workbook = load_workbook(path)
    workbook.active["C2"] = "not-an-email"
    workbook.save(path)
    session = AsyncMock()
    result = await import_customers(path, session, apply=False)
    assert not result.ok
    assert "invalid email" in result.errors[0]["errors"][0]


@pytest.mark.asyncio
async def test_invalid_customer_csv_email_is_row_error(tmp_path: Path) -> None:
    path = tmp_path / "customers.csv"
    path.write_text(
        "company_name,contact_name,email,language,product_code,currency,auto_send_allowed,consent_basis,do_not_contact\n"
        "Demo,Buyer,not-an-email,en,WIDGET-100,USD,true,existing relationship,false\n",
        encoding="utf-8",
    )
    session = AsyncMock()
    result = await import_customers(path, session, apply=False)
    assert not result.ok
    assert "invalid email" in result.errors[0]["errors"][0]


@pytest.mark.asyncio
async def test_missing_product_is_valid_contact_only_row(tmp_path: Path) -> None:
    generate_templates(tmp_path)
    path = tmp_path / "customer_list_template.xlsx"
    workbook = load_workbook(path)
    workbook.active["E2"] = "UNKNOWN-PRODUCT"
    workbook.save(path)
    session = AsyncMock()
    session.scalar.return_value = None

    result = await import_customers(path, session, apply=False)

    assert result.ok
    assert result.valid_rows == 1
    assert result.contact_only_rows == 1
    assert result.case_ready_rows == 0
    assert result.missing_product_codes == ["UNKNOWN-PRODUCT"]


@pytest.mark.integration
@pytest.mark.asyncio
async def test_contact_only_customer_import_is_applied_idempotently(
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    generate_templates(tmp_path)
    path = tmp_path / "customer_list_template.xlsx"
    workbook = load_workbook(path)
    workbook.active["E2"] = "UNKNOWN-PRODUCT"
    workbook.save(path)

    first = await import_customers(path, db_session, apply=True)
    second = await import_customers(path, db_session, apply=True)

    assert first.ok and second.ok
    assert first.created_customers == 1
    assert first.created_contacts == 1
    assert first.created_cases == 0
    assert second.created_customers == 0
    assert second.created_contacts == 0
    assert await db_session.scalar(select(func.count()).select_from(Customer)) == 1
    assert await db_session.scalar(select(func.count()).select_from(Contact)) == 1
    assert await db_session.scalar(select(func.count()).select_from(SalesCase)) == 0


@pytest.mark.asyncio
async def test_floor_above_standard_blocks_price_import(tmp_path: Path) -> None:
    generate_templates(tmp_path)
    path = tmp_path / "price_list_template.xlsx"
    workbook = load_workbook(path)
    workbook.active["G2"] = "50"
    workbook.active["H2"] = "60"
    workbook.save(path)
    session = AsyncMock()
    session.scalar.return_value = None
    result = await import_prices(path, session, apply=False)
    assert not result.ok
    assert any("floor cannot exceed" in error for error in result.errors[0]["errors"])


@pytest.mark.asyncio
async def test_manual_only_product_allows_blank_prices(tmp_path: Path) -> None:
    generate_templates(tmp_path)
    path = tmp_path / "price_list_template.xlsx"
    workbook = load_workbook(path)
    sheet = workbook.active
    columns = {cell.value: cell.column for cell in sheet[1]}
    sheet.cell(2, columns["product_code"], "YAC-TBDMSC")
    sheet.cell(2, columns["product_name"], "YAC-TBDMSC")
    sheet.cell(2, columns["approved_text_key"], "yac_tbdmsc")
    sheet.cell(2, columns["margin_class"], "A")
    sheet.cell(2, columns["currency"], "INR")
    sheet.cell(2, columns["unit"], "kg")
    sheet.cell(2, columns["standard_price"], None)
    sheet.cell(2, columns["absolute_floor"], None)
    sheet.cell(2, columns["manual_only"], True)
    workbook.save(path)
    session = AsyncMock()
    session.scalar.return_value = None

    result = await import_prices(path, session, apply=False)

    assert result.ok
    assert result.valid_rows == 1
