from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from email import policy
from email.message import EmailMessage
from email.parser import BytesParser
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.ai import AIClient, CompanyCategoryDecision, CompanyResearchSource
from app.db import (
    AIInvocation,
    CaseStage,
    CaseStatus,
    Contact,
    Customer,
    DeliveryStatus,
    Handoff,
    Outbox,
    PricePolicy,
    Product,
    ProductCategory,
    ReactivationCampaign,
    ReactivationRecipient,
    SalesCase,
)
from app.db import (
    EmailMessage as DBEmailMessage,
)
from app.domain import HandoffReason
from app.product_catalog import import_product_catalog
from app.services import (
    backfill_product_list_requests,
    ingest_raw_email,
    process_inbound,
    send_one_outbox,
)
from app.settings import get_settings

pytestmark = pytest.mark.integration


async def _seed_catalog_and_interest(
    db_session: AsyncSession,
    *,
    interests: list[str],
    auto_send_allowed: bool = True,
) -> tuple[int, int]:
    await import_product_catalog(db_session, apply=True)
    customer = Customer(
        company_name="Ethachem",
        language="en",
        auto_send_allowed=auto_send_allowed,
        consent_basis="existing CRM/customer-list relationship",
        metadata_json={},
    )
    db_session.add(customer)
    await db_session.flush()
    if interests:
        from app.product_catalog import category_names_by_key, interest_entry, merge_customer_interests

        names = await category_names_by_key(db_session)
        merge_customer_interests(
            customer,
            [
                interest_entry(
                    category_key=key,
                    category_name=names.get(key, key),
                    source="test",
                    value=key,
                )
                for key in interests
            ],
        )
    contact = Contact(
        customer_id=customer.id,
        name="Alice Buyer",
        email="api@ethachem.example",
        language="en",
    )
    db_session.add(contact)
    await db_session.commit()
    return customer.id, contact.id


async def _seed_departed_reactivation_parent(
    db_session: AsyncSession,
    *,
    interests: list[str],
    old_email: str = "globalsourcing@witofly.com",
    reply_email: str = "marketing001@witofly.com",
) -> tuple[int, int, Outbox]:
    await import_product_catalog(db_session, apply=True)
    customer = Customer(
        company_name="Shanghai Witofly Chemical Co.,Ltd",
        language="en",
        auto_send_allowed=True,
        consent_basis="existing CRM/customer-list relationship",
        metadata_json={},
    )
    db_session.add(customer)
    await db_session.flush()
    if interests:
        from app.product_catalog import (
            category_names_by_key,
            interest_entry,
            merge_customer_interests,
        )

        names = await category_names_by_key(db_session)
        merge_customer_interests(
            customer,
            [
                interest_entry(
                    category_key=key,
                    category_name=names.get(key, key),
                    source="test",
                    value=key,
                )
                for key in interests
            ],
        )
    old_contact = Contact(
        customer_id=customer.id,
        name="Pooja Raut",
        email=old_email,
        language="en",
    )
    db_session.add(old_contact)
    await db_session.flush()
    parent = Outbox(
        case_id=None,
        quote_id=None,
        message_kind="REACTIVATION",
        business_key="reactivation:test:departed",
        message_id="<wake-departed@lanyachemindia.com>",
        recipient=old_email,
        raw_message=f"From: sales-agent@example.com\nTo: {old_email}\n\nChecking in",
        status=DeliveryStatus.SENT,
        sent_at=datetime.now(UTC) - timedelta(minutes=10),
        sent_via="smtp",
    )
    db_session.add(parent)
    await db_session.flush()
    campaign = ReactivationCampaign(
        name="Departed reply test",
        status="RUNNING",
        subject_template="Checking in",
        body_template="Hello",
        start_date=date.today(),
        created_by="test",
    )
    db_session.add(campaign)
    await db_session.flush()
    recipient = ReactivationRecipient(
        campaign_id=campaign.id,
        customer_id=customer.id,
        contact_id=old_contact.id,
        outbox_id=parent.id,
        status="SENT",
        eligible=True,
        selected=True,
        sent_at=parent.sent_at,
    )
    db_session.add(recipient)
    outbound = DBEmailMessage(
        case_id=None,
        customer_id=customer.id,
        contact_id=old_contact.id,
        direction="OUTBOUND",
        message_id=parent.message_id,
        from_address="sales-agent@example.com",
        to_addresses=[old_email],
        subject="Checking in from Lanya Chem",
        body_text="Checking in",
        raw_sha256="d" * 64,
        received_at=parent.sent_at,
    )
    db_session.add_all([recipient, outbound])
    await db_session.commit()
    return customer.id, old_contact.id, parent


def _departed_reply_message(parent_message_id: str, body: str) -> bytes:
    message = EmailMessage()
    message["From"] = "Judy Ao <marketing001@witofly.com>"
    message["To"] = "sales-agent@example.com"
    message["Subject"] = "Re: Checking in from Lanya Chem"
    message["Message-ID"] = "<departed-reply@example.com>"
    message["In-Reply-To"] = parent_message_id
    message["References"] = parent_message_id
    message.set_content(body)
    return message.as_bytes()


def _message(
    subject: str,
    body: str,
    message_id: str = "product-list@example.com",
    *,
    in_reply_to: str | None = None,
    references: str | None = None,
) -> bytes:
    message = EmailMessage()
    message["From"] = "Alice Buyer <api@ethachem.example>"
    message["To"] = "sales-agent@example.com"
    message["Subject"] = subject
    message["Message-ID"] = f"<{message_id}>"
    if in_reply_to is not None:
        message["In-Reply-To"] = in_reply_to
    if references is not None:
        message["References"] = references
    message.set_content(body)
    return message.as_bytes()


async def _queued_product_list(db_session: AsyncSession) -> Outbox | None:
    return await db_session.scalar(
        select(Outbox).where(Outbox.message_kind == "PRODUCT_LIST")
    )


async def test_crm_interest_triggers_automatic_product_list_reply(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product list inquiry",
            "Please send us your product list for industrial silane.",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    case = await db_session.get(SalesCase, email_row.case_id)
    assert case is not None
    assert case.product_id is None
    assert case.category_id is not None
    category = await db_session.get(ProductCategory, case.category_id)
    assert category is not None and category.key == "industrial_silanes"
    assert await db_session.scalar(select(func.count()).select_from(Handoff)) == 0

    await process_inbound(db_session, email_row.id)

    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    assert outbox.business_key == f"inbound-product-list:{email_row.id}"
    assert "YAC-A110" in outbox.raw_message
    assert "919-30-2" in outbox.raw_message
    assert "USD" not in outbox.raw_message
    assert "Please send us your product list for industrial silane." in outbox.raw_message
    # One copy in the plain part and one in the HTML part; neither part should
    # duplicate the sign-off already supplied by the configured signature.
    assert outbox.raw_message.count("Best regards,") == 2
    assert await db_session.scalar(select(func.count()).select_from(Handoff)) == 0

    assert await send_one_outbox(db_session) is True
    await db_session.refresh(outbox)
    assert outbox.status == DeliveryStatus.SENT
    outbox_dir = get_settings().runtime_dir / "demo_outbox"
    assert any(outbox_dir.glob("*.eml"))


async def test_excel_cas_request_attaches_verified_catalog_workbook(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["pharmaceutical"])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product data request",
            "Please share your product with CAS# in excel sheet.",
            message_id="excel-cas-list@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    await process_inbound(db_session, email_row.id)

    outbox = await _queued_product_list(db_session)
    if outbox is None:
        invocation = await db_session.scalar(
            select(AIInvocation)
            .where(AIInvocation.case_id == email_row.case_id)
            .order_by(AIInvocation.id.desc())
        )
        handoffs = (
            (
                await db_session.execute(
                    select(Handoff).where(Handoff.case_id == email_row.case_id)
                )
            )
            .scalars()
            .all()
        )
        pytest.fail(
            "catalog workbook was not queued: "
            f"analysis={invocation.parsed_output if invocation else None}; "
            f"handoffs={[(item.reason_code, item.summary) for item in handoffs]}"
        )
    assert outbox.business_key == f"inbound-product-list:{email_row.id}"
    mime = BytesParser(policy=policy.default).parsebytes(
        outbox.raw_message.encode("utf-8")
    )
    attachments = [
        part
        for part in mime.walk()
        if part.get_content_disposition() == "attachment"
    ]
    assert len(attachments) == 1
    attachment = attachments[0]
    assert attachment.get_filename() == (
        "Lanya_Chem_pharmaceutical_product_list.xlsx"
    )
    workbook = load_workbook(
        BytesIO(attachment.get_payload(decode=True)),
        data_only=False,
    )
    sheet = workbook["Product List"]
    acac_row = next(
        row
        for row in sheet.iter_rows(min_row=2, values_only=True)
        if row[2] == "ACAC"
    )
    assert acac_row[4] is None
    outbound_email = await db_session.scalar(
        select(DBEmailMessage).where(
            DBEmailMessage.direction == "OUTBOUND",
            DBEmailMessage.message_id == outbox.message_id,
        )
    )
    assert outbound_email is not None
    assert any(
        item["filename"] == "Lanya_Chem_pharmaceutical_product_list.xlsx"
        for item in outbound_email.attachment_metadata
    ), outbound_email.attachment_metadata


async def test_generic_category_interest_email_auto_replies(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Hello",
            "We are interested in industrial silane.",
            message_id="generic-interest@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    await process_inbound(db_session, email_row.id)

    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    assert "YAC-S313" in outbox.raw_message


async def test_productless_quote_with_category_interest_sends_one_clarification(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Quotation request",
            "Please quote quantity: 100 kg.",
            message_id="productless-quote@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    await process_inbound(db_session, email_row.id)

    clarification = await db_session.scalar(
        select(Outbox).where(Outbox.message_kind == "QUOTE_CLARIFICATION")
    )
    assert clarification is not None
    assert clarification.business_key == (
        f"inbound-reply:{email_row.id}:clarification"
    )
    assert "quotation request for 100 kg" in clarification.raw_message
    assert "confirm the product name or Lanya product code" in clarification.raw_message
    assert "Please quote quantity: 100 kg." in clarification.raw_message
    mime = BytesParser(policy=policy.default).parsebytes(
        clarification.raw_message.encode("utf-8")
    )
    assert mime["In-Reply-To"] == email_row.message_id
    assert await db_session.scalar(select(func.count()).select_from(Handoff)) == 0


async def test_productless_quote_recovers_product_from_complete_quoted_thread(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Re: Our requirement",
            (
                "Please quote 100 kg.\n\n"
                "On Monday Alice Buyer wrote:\n"
                "> We are interested in YAC-A110."
            ),
            message_id="quoted-product@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    assert "YAC-A110" not in email_row.body_text
    await process_inbound(db_session, email_row.id)

    case = await db_session.get(SalesCase, email_row.case_id)
    assert case is not None
    await db_session.refresh(case, ["product"])
    assert case.product is not None and case.product.code == "YAC-A110"
    assert await db_session.scalar(
        select(func.count())
        .select_from(Outbox)
        .where(Outbox.message_kind == "QUOTE_CLARIFICATION")
    ) == 0


async def test_second_productless_quote_after_clarification_routes_to_human(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    first = await ingest_raw_email(
        db_session,
        _message(
            "Quotation request",
            "Please quote quantity: 100 kg.",
            message_id="clarification-first@example.com",
        ),
        mailbox="integration-test",
    )
    assert first is not None and first.case_id is not None
    await process_inbound(db_session, first.id)
    clarification = await db_session.scalar(
        select(Outbox).where(Outbox.message_kind == "QUOTE_CLARIFICATION")
    )
    assert clarification is not None

    second = await ingest_raw_email(
        db_session,
        _message(
            f"Re: {first.subject}",
            "Yes, please quote 100 kg.",
            message_id="clarification-second@example.com",
            in_reply_to=clarification.message_id,
            references=clarification.message_id,
        ),
        mailbox="integration-test",
    )
    assert second is not None and second.case_id == first.case_id
    await process_inbound(db_session, second.id)

    assert await db_session.scalar(
        select(func.count())
        .select_from(Outbox)
        .where(Outbox.message_kind == "QUOTE_CLARIFICATION")
    ) == 1
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == second.id)
    )
    assert handoff is not None
    assert handoff.reason_code == HandoffReason.HUMAN_CONTROL.value
    assert "still unclear" in handoff.summary


async def test_sample_request_still_requires_human(db_session: AsyncSession) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Sample request",
            "Please send a sample of your industrial silane products.",
            message_id="sample@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    await process_inbound(db_session, email_row.id)

    assert await _queued_product_list(db_session) is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None
    assert handoff.reason_code == HandoffReason.SAMPLE_REQUEST.value


async def test_unknown_interest_routes_to_semantic_handoff_when_research_disabled(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=[])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product list inquiry",
            "Please send us your product list.",
            message_id="no-interest@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is None

    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None
    assert handoff.reason_code == HandoffReason.PRODUCT_CATEGORY_REVIEW.value
    assert handoff.extracted_facts["company_research"]["status"] == "DISABLED"
    assert await _queued_product_list(db_session) is None


async def test_company_research_observation_mode_records_evidence_without_sending(
    db_session: AsyncSession,
    monkeypatch,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=[])
    settings = get_settings()
    monkeypatch.setattr(settings, "company_research_enabled", True)
    monkeypatch.setattr(settings, "company_research_auto_send_enabled", False)

    async def research(*args, **kwargs):
        return (
            CompanyCategoryDecision(
                identity_confidence=0.98,
                recommended_category_key="industrial_silanes",
                category_confidence=0.94,
                runner_up_category_key="rubber_plastics",
                runner_up_confidence=0.20,
                conflicting_evidence=False,
                rationale="Two sources identify industrial silane distribution.",
            ),
            [
                CompanyResearchSource(
                    url="https://industry.example/ethachem",
                    title="Industry directory",
                    cited_text="Industrial silane distributor",
                ),
                CompanyResearchSource(
                    url="https://trade.example/ethachem",
                    title="Trade profile",
                    cited_text="Silane coupling agents",
                ),
            ],
            {
                "provider": "anthropic",
                "model": "claude-test",
                "request_hash": "a" * 64,
                "input_tokens": 20,
                "output_tokens": 10,
            },
        )

    monkeypatch.setattr(AIClient, "research_company_category", research)
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product list inquiry",
            "Please send us your product list.",
            message_id="research-observation@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None and email_row.case_id is not None

    await process_inbound(db_session, email_row.id)

    assert await _queued_product_list(db_session) is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None
    assert handoff.reason_code == HandoffReason.PRODUCT_CATEGORY_REVIEW.value
    research_facts = handoff.extracted_facts["company_research"]
    assert research_facts["gate"]["eligible"] is True
    assert research_facts["decision"]["recommended_category_key"] == "industrial_silanes"
    invocation = await db_session.scalar(
        select(AIInvocation).where(AIInvocation.purpose == "company_category_research")
    )
    assert invocation is not None and invocation.success is True


async def test_company_research_high_confidence_auto_sends_and_reuses_cache(
    db_session: AsyncSession,
    monkeypatch,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=[])
    settings = get_settings()
    monkeypatch.setattr(settings, "company_research_enabled", True)
    monkeypatch.setattr(settings, "company_research_auto_send_enabled", True)
    calls = 0

    async def research(*args, **kwargs):
        nonlocal calls
        calls += 1
        return (
            CompanyCategoryDecision(
                identity_confidence=0.98,
                recommended_category_key="industrial_silanes",
                category_confidence=0.94,
                runner_up_category_key="rubber_plastics",
                runner_up_confidence=0.20,
                conflicting_evidence=False,
                rationale="Two sources identify industrial silane distribution.",
            ),
            [
                CompanyResearchSource(url="https://industry.example/ethachem"),
                CompanyResearchSource(url="https://trade.example/ethachem"),
            ],
            {
                "provider": "anthropic",
                "model": "claude-test",
                "request_hash": "b" * 64,
            },
        )

    monkeypatch.setattr(AIClient, "research_company_category", research)
    first = await ingest_raw_email(
        db_session,
        _message(
            "First catalog request",
            "Please send us your product list.",
            message_id="research-auto-first@example.com",
        ),
        mailbox="integration-test",
    )
    assert first is not None and first.case_id is not None
    await process_inbound(db_session, first.id)

    first_outbox = await db_session.scalar(
        select(Outbox).where(Outbox.business_key == f"inbound-product-list:{first.id}")
    )
    assert first_outbox is not None
    first_case = await db_session.get(SalesCase, first.case_id)
    assert first_case is not None and first_case.category_id is not None

    second = await ingest_raw_email(
        db_session,
        _message(
            "Second catalog request",
            "Please send your product catalog.",
            message_id="research-auto-second@example.com",
        ),
        mailbox="integration-test",
    )
    assert second is not None and second.case_id is not None
    await process_inbound(db_session, second.id)

    second_outbox = await db_session.scalar(
        select(Outbox).where(Outbox.business_key == f"inbound-product-list:{second.id}")
    )
    assert second_outbox is not None
    assert calls == 1


async def test_selected_legacy_handoff_can_use_company_research_backfill(
    db_session: AsyncSession,
    monkeypatch,
) -> None:
    customer_id, contact_id = await _seed_catalog_and_interest(db_session, interests=[])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Legacy product list request",
            "Please send us your product list.",
            message_id="research-backfill@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None and email_row.case_id is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None

    sales_case = SalesCase(
        customer_id=customer_id,
        contact_id=contact_id,
        product_id=None,
        category_id=None,
        currency="INR",
        stage=CaseStage.QUOTING,
        status=CaseStatus.WAITING_HUMAN,
        subject_key="legacy product list request",
    )
    db_session.add(sales_case)
    await db_session.flush()
    email_row.case_id = sales_case.id
    handoff.case_id = sales_case.id
    await db_session.commit()

    settings = get_settings()
    monkeypatch.setattr(settings, "company_research_enabled", True)
    monkeypatch.setattr(settings, "company_research_auto_send_enabled", True)
    calls = 0

    async def research(*args, **kwargs):
        nonlocal calls
        calls += 1
        return (
            CompanyCategoryDecision(
                identity_confidence=0.98,
                recommended_category_key="industrial_silanes",
                category_confidence=0.94,
                runner_up_category_key="rubber_plastics",
                runner_up_confidence=0.20,
                conflicting_evidence=False,
                rationale="Two sources identify industrial silane distribution.",
            ),
            [
                CompanyResearchSource(url="https://industry.example/ethachem"),
                CompanyResearchSource(url="https://trade.example/ethachem"),
            ],
            {
                "provider": "anthropic",
                "model": "claude-test",
                "request_hash": "c" * 64,
            },
        )

    monkeypatch.setattr(AIClient, "research_company_category", research)
    preview = await backfill_product_list_requests(
        db_session,
        apply=False,
        handoff_ids=(handoff.id,),
        company_research=True,
    )
    assert preview["candidate_count"] == 1
    assert preview["candidates"][0]["company_research_required"] is True
    assert calls == 0

    result = await backfill_product_list_requests(
        db_session,
        apply=True,
        handoff_ids=(handoff.id,),
        company_research=True,
    )
    assert result["queued_count"] == 1
    assert calls == 1
    outbox = await db_session.scalar(
        select(Outbox).where(
            Outbox.business_key == f"inbound-product-list:{email_row.id}"
        )
    )
    assert outbox is not None
    await db_session.refresh(handoff)
    await db_session.refresh(sales_case)
    assert handoff.status == "RESOLVED"
    assert sales_case.status == CaseStatus.ACTIVE
    assert sales_case.category_id is not None


async def test_multiple_interests_route_to_human(db_session: AsyncSession) -> None:
    await _seed_catalog_and_interest(
        db_session,
        interests=["industrial_silanes", "pharmaceutical"],
    )
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product list inquiry",
            "Please send us your product list.",
            message_id="multi-interest@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None
    assert handoff.reason_code == HandoffReason.NEW_INQUIRY_REVIEW.value
    assert handoff.extracted_facts["active_interest_categories"] == [
        "industrial_silanes",
        "pharmaceutical",
    ]


async def test_product_list_backfill_previews_then_queues_old_open_handoff(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product catalog",
            "Please send your product list.",
            message_id="backfill-product-list@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None and email_row.case_id is not None
    case = await db_session.get(SalesCase, email_row.case_id)
    assert case is not None
    case.status = CaseStatus.WAITING_HUMAN
    handoff = Handoff(
        case_id=case.id,
        source_email_id=email_row.id,
        reason_code=HandoffReason.HUMAN_CONTROL.value,
        summary="Legacy product-list request requires review",
        extracted_facts={"product_pending": True},
        status="OPEN",
        dingtalk_status="SENT",
    )
    db_session.add(handoff)
    await db_session.commit()

    preview = await backfill_product_list_requests(db_session, apply=False)
    assert preview["candidate_count"] == 1
    assert preview["queued_count"] == 0
    assert preview["candidates"][0]["category_key"] == "industrial_silanes"
    assert await _queued_product_list(db_session) is None
    await db_session.refresh(handoff)
    assert handoff.status == "OPEN"

    result = await backfill_product_list_requests(
        db_session,
        apply=True,
        handoff_ids=(handoff.id,),
    )
    assert result["candidate_count"] == 1
    assert result["queued_count"] == 1
    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    assert "YAC-A110" in outbox.raw_message
    await db_session.refresh(handoff)
    await db_session.refresh(case)
    assert handoff.status == "RESOLVED"
    assert case.status == CaseStatus.ACTIVE


async def test_product_list_backfill_creates_case_after_interest_is_mapped(
    db_session: AsyncSession,
) -> None:
    customer_id, _ = await _seed_catalog_and_interest(db_session, interests=[])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product catalog",
            "Please send your product list.",
            message_id="backfill-new-category-case@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None and email_row.case_id is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None and handoff.case_id is None

    from app.product_catalog import category_names_by_key, interest_entry, merge_customer_interests

    customer = await db_session.get(Customer, customer_id)
    assert customer is not None
    names = await category_names_by_key(db_session)
    merge_customer_interests(
        customer,
        [
            interest_entry(
                category_key="industrial_silanes",
                category_name=names["industrial_silanes"],
                source="test-backfill",
                value="industrial_silanes",
            )
        ],
    )
    await db_session.commit()

    result = await backfill_product_list_requests(
        db_session,
        apply=True,
        handoff_ids=(handoff.id,),
    )
    assert result["queued_count"] == 1
    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    case = await db_session.get(SalesCase, outbox.case_id)
    assert case is not None
    assert case.category_id is not None
    assert case.product_id is None
    await db_session.refresh(handoff)
    assert handoff.case_id == case.id
    assert handoff.status == "RESOLVED"


async def test_product_list_backfill_maps_specific_product_in_selected_category(
    db_session: AsyncSession,
) -> None:
    customer_id, _ = await _seed_catalog_and_interest(db_session, interests=[])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product catalog",
            "Please send your product list for ACAC.",
            message_id="backfill-specific-product@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None and email_row.case_id is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None

    from app.product_catalog import category_names_by_key, interest_entry, merge_customer_interests

    customer = await db_session.get(Customer, customer_id)
    assert customer is not None
    names = await category_names_by_key(db_session)
    merge_customer_interests(
        customer,
        [
            interest_entry(
                category_key="pharmaceutical",
                category_name=names["pharmaceutical"],
                source="test-backfill",
                value="Acetyl Acetone",
            )
        ],
    )
    await db_session.commit()

    preview = await backfill_product_list_requests(
        db_session,
        apply=False,
        handoff_ids=(handoff.id,),
    )
    assert preview["candidate_count"] == 1
    assert preview["candidates"][0]["detected_product_code"] == "ACAC"
    assert preview["candidates"][0]["matched_product_id"] is not None

    result = await backfill_product_list_requests(
        db_session,
        apply=True,
        handoff_ids=(handoff.id,),
    )
    assert result["queued_count"] == 1
    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    case = await db_session.scalar(
        select(SalesCase)
        .options(selectinload(SalesCase.product))
        .where(SalesCase.id == outbox.case_id)
    )
    assert case is not None and case.product is not None
    assert case.product.code == "ACAC"


async def test_product_list_backfill_preflight_failure_does_not_create_case(
    db_session: AsyncSession,
) -> None:
    customer_id, _ = await _seed_catalog_and_interest(db_session, interests=[])
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product catalog",
            "Please send your product list.",
            message_id="backfill-missing-mime@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None and email_row.case_id is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None and handoff.case_id is None

    from app.product_catalog import category_names_by_key, interest_entry, merge_customer_interests

    customer = await db_session.get(Customer, customer_id)
    assert customer is not None
    names = await category_names_by_key(db_session)
    merge_customer_interests(
        customer,
        [
            interest_entry(
                category_key="industrial_silanes",
                category_name=names["industrial_silanes"],
                source="test-backfill",
                value="industrial_silanes",
            )
        ],
    )
    email_row.body_html = '<p>Please send your product list.</p><img src="cid:missing-logo">'
    archive_path = (
        get_settings().runtime_dir
        / "inbound_archive"
        / f"{email_row.raw_sha256}.eml"
    )
    archive_path.unlink()
    await db_session.commit()

    preview = await backfill_product_list_requests(
        db_session,
        apply=False,
        handoff_ids=(handoff.id,),
    )
    assert preview["candidate_count"] == 0
    assert preview["exclusion_counts"] == {"REPLY_SOURCE_OR_RENDER_UNAVAILABLE": 1}

    result = await backfill_product_list_requests(
        db_session,
        apply=True,
        handoff_ids=(handoff.id,),
    )
    assert result["candidate_count"] == 0
    assert result["queued_count"] == 0
    assert await _queued_product_list(db_session) is None
    await db_session.refresh(email_row)
    await db_session.refresh(handoff)
    assert email_row.case_id is None
    assert handoff.case_id is None
    assert handoff.status == "OPEN"


async def test_product_list_backfill_excludes_non_unique_excel_interests(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(
        db_session,
        interests=["industrial_silanes", "pharmaceutical"],
    )
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product catalog",
            "Please send your product list.",
            message_id="backfill-ambiguous@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None

    result = await backfill_product_list_requests(db_session, apply=False)
    assert result["candidate_count"] == 0
    assert result["exclusion_counts"]["INTEREST_CATEGORY_NOT_UNIQUE"] == 1
    assert await _queued_product_list(db_session) is None


async def test_product_specific_list_request_sends_product_category(
    db_session: AsyncSession,
) -> None:
    await _seed_catalog_and_interest(db_session, interests=["industrial_silanes"])
    product = await db_session.scalar(
        select(Product).where(Product.code == "YAC-A110")
    )
    assert product is not None
    db_session.add(
        PricePolicy(
            product_id=product.id,
            currency="USD",
            standard_price=Decimal("10.0000"),
            absolute_floor=Decimal("8.0000"),
            max_discount_pct=Decimal("0.1000"),
            max_negotiation_rounds=2,
            concession_step_pct=Decimal("0.0200"),
            min_quantity=1,
            max_quantity=100000,
            quote_valid_days=30,
            standard_incoterm="EXW",
            allowed_incoterms=["EXW"],
            standard_payment_term="100% before shipment",
            allowed_payment_terms=["100% before shipment"],
            valid_from=date.today(),
            source_hash="test-product-list",
        )
    )
    await db_session.commit()
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "YAC-A110 product list",
            "PRODUCT YAC-A110. Please send your product list.",
            message_id="a110-list@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    case = await db_session.get(SalesCase, email_row.case_id)
    assert case is not None and case.product_id is not None
    product = await db_session.get(Product, case.product_id)
    assert product is not None and product.code == "YAC-A110"

    await process_inbound(db_session, email_row.id)

    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    assert "YAC-A110" in outbox.raw_message
    assert "YAC-N113" in outbox.raw_message


async def test_catalog_import_is_idempotent(db_session: AsyncSession) -> None:
    first = await import_product_catalog(db_session, apply=True)
    mtms = await db_session.scalar(select(Product).where(Product.code == "YAC-MTMS"))
    assert mtms is not None
    mtms.cas_no = "1185-55-3"
    await db_session.flush()
    second = await import_product_catalog(db_session, apply=True)

    assert first["products_created"] == 71
    assert first["categories_created"] == 3
    assert second["products_created"] == 0
    assert second["products_updated"] == 71
    product_count = await db_session.scalar(select(func.count()).select_from(Product))
    category_count = await db_session.scalar(
        select(func.count()).select_from(ProductCategory)
    )
    assert product_count == 71
    assert category_count == 3
    assert mtms.cas_no is None


async def test_suppressed_customer_never_auto_sends(db_session: AsyncSession) -> None:
    await _seed_catalog_and_interest(
        db_session,
        interests=["industrial_silanes"],
        auto_send_allowed=False,
    )
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product list inquiry",
            "Please send us your product list for industrial silane.",
            message_id="no-auto-send@example.com",
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    await process_inbound(db_session, email_row.id)

    assert await _queued_product_list(db_session) is None
    handoff = await db_session.scalar(
        select(Handoff).where(Handoff.source_email_id == email_row.id)
    )
    assert handoff is not None
    assert handoff.reason_code == HandoffReason.SUPPRESSED.value


async def test_full_customer_workbook_stores_interest_category(
    db_session: AsyncSession,
    tmp_path,
) -> None:
    from datetime import date

    from openpyxl import Workbook

    from app.full_customer_import import (
        COMPANY_HEADER,
        CONTACT_HEADER,
        EMAIL_HEADER,
        FIRST_CONTACT_HEADER,
        LAST_CONTACT_HEADER,
        NO_AI_HEADER,
        OTHER_EMAIL_HEADER,
        PRODUCT_HEADER,
        import_full_customer_workbook,
    )

    await import_product_catalog(db_session, apply=True)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "customers"
    sheet.append(
        [
            COMPANY_HEADER,
            CONTACT_HEADER,
            EMAIL_HEADER,
            OTHER_EMAIL_HEADER,
            PRODUCT_HEADER,
            FIRST_CONTACT_HEADER,
            LAST_CONTACT_HEADER,
            NO_AI_HEADER,
        ]
    )
    sheet.append(
        [
            "Ethachem",
            "Alice Buyer",
            "api@ethachem.example",
            "",
            "工业硅烷",
            date(2020, 1, 1),
            date(2024, 1, 1),
            "",
        ]
    )
    sheet.append(
        [
            "Brisben CHEM",
            "Bijesh Shah",
            "brisbenchem@example.com",
            "",
            "Acetyl Acetone",
            date(2023, 9, 22),
            date(2023, 9, 22),
            "",
        ]
    )
    path = tmp_path / "customers.xlsx"
    workbook.save(path)

    await import_full_customer_workbook(
        path,
        db_session,
        apply=True,
        enable_auto_send=True,
    )

    customer = await db_session.scalar(
        select(Customer).where(Customer.company_name == "Ethachem")
    )
    assert customer is not None
    interests = (customer.metadata_json or {}).get("interests")
    assert interests is not None
    assert interests[0]["category_key"] == "industrial_silanes"
    assert interests[0]["value"] == "工业硅烷"
    assert customer.auto_send_allowed is True

    brisben = await db_session.scalar(
        select(Customer).where(Customer.company_name == "Brisben CHEM")
    )
    assert brisben is not None
    brisben_interests = (brisben.metadata_json or {}).get("interests")
    assert brisben_interests is not None
    assert brisben_interests[0]["category_key"] == "pharmaceutical"
    assert brisben_interests[0]["value"] == "Acetyl Acetone"
    brisben_contact = await db_session.scalar(
        select(Contact).where(Contact.email == "brisbenchem@example.com")
    )
    acac = await db_session.scalar(select(Product).where(Product.code == "ACAC"))
    assert brisben_contact is not None
    assert acac is not None and acac.category_id is not None
    acac_case = await db_session.scalar(
        select(SalesCase).where(
            SalesCase.contact_id == brisben_contact.id,
            SalesCase.product_id == acac.id,
        )
    )
    assert acac_case is not None
    assert acac_case.category_id == acac.category_id

    # The stored interest now drives an automatic product-list reply.
    contact = await db_session.scalar(
        select(Contact).where(Contact.email == "api@ethachem.example")
    )
    assert contact is not None
    email_row = await ingest_raw_email(
        db_session,
        _message(
            "Product list inquiry",
            "Please send us your product list for industrial silane.",
            message_id="workbook-import@example.com",
        ),
        mailbox="integration-test",
    )
    assert email_row is not None and email_row.case_id is not None
    await process_inbound(db_session, email_row.id)
    outbox = await _queued_product_list(db_session)
    assert outbox is not None


async def test_departed_reply_from_same_domain_retires_old_contact_and_auto_sends(
    db_session: AsyncSession,
) -> None:
    customer_id, old_contact_id, parent = await _seed_departed_reactivation_parent(
        db_session,
        interests=["industrial_silanes"],
    )
    email_row = await ingest_raw_email(
        db_session,
        _departed_reply_message(
            parent.message_id,
            (
                "Dear Shreya\n\n"
                "Ms. Pooja no longer works in our company.\n\n"
                "Please send us your product list.\n\n"
                "Regards, Judy"
            ),
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    assert email_row.automated_reply_type == "DEPARTED"
    assert email_row.automated_reply_metadata["personnel_change_handled"] is True

    old_contact = await db_session.get(Contact, old_contact_id)
    assert old_contact is not None and old_contact.suppressed is True

    new_contact = await db_session.scalar(
        select(Contact).where(
            func.lower(Contact.email) == "marketing001@witofly.com"
        )
    )
    assert new_contact is not None and new_contact.suppressed is False
    assert new_contact.customer_id == customer_id

    case = await db_session.get(SalesCase, email_row.case_id)
    assert case is not None and case.category_id is not None
    category = await db_session.get(ProductCategory, case.category_id)
    assert category is not None and category.key == "industrial_silanes"

    await process_inbound(db_session, email_row.id)

    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    assert "marketing001@witofly.com" in outbox.raw_message
    assert "YAC-A110" in outbox.raw_message
    assert await db_session.scalar(
        select(func.count()).select_from(Handoff)
    ) == 0
    recipient = await db_session.scalar(
        select(ReactivationRecipient).where(
            ReactivationRecipient.outbox_id == parent.id
        )
    )
    assert recipient is not None and recipient.status == "REPLIED"


async def test_departed_reply_without_interest_researches_and_auto_sends(
    db_session: AsyncSession,
    monkeypatch,
) -> None:
    customer_id, old_contact_id, parent = await _seed_departed_reactivation_parent(
        db_session,
        interests=[],
    )
    settings = get_settings()
    monkeypatch.setattr(settings, "company_research_enabled", True)
    monkeypatch.setattr(settings, "company_research_auto_send_enabled", True)

    async def research(*args, **kwargs):
        return (
            CompanyCategoryDecision(
                identity_confidence=0.98,
                recommended_category_key="industrial_silanes",
                category_confidence=0.94,
                runner_up_category_key="rubber_plastics",
                runner_up_confidence=0.20,
                conflicting_evidence=False,
                rationale="Two sources identify industrial silane distribution.",
            ),
            [
                CompanyResearchSource(
                    url="https://industry.example/witofly",
                    title="Industry directory",
                    cited_text="Industrial silane distributor",
                ),
                CompanyResearchSource(
                    url="https://trade.example/witofly",
                    title="Trade profile",
                    cited_text="Silane coupling agents",
                ),
            ],
            {
                "provider": "anthropic",
                "model": "claude-test",
                "request_hash": "e" * 64,
                "input_tokens": 20,
                "output_tokens": 10,
            },
        )

    monkeypatch.setattr(AIClient, "research_company_category", research)
    email_row = await ingest_raw_email(
        db_session,
        _departed_reply_message(
            parent.message_id,
            (
                "Dear Shreya\n\n"
                "Ms. Pooja no longer works in our company.\n\n"
                "Please send us your product list.\n\n"
                "Regards, Judy"
            ),
        ),
        mailbox="integration-test",
    )

    assert email_row is not None and email_row.case_id is not None
    old_contact = await db_session.get(Contact, old_contact_id)
    assert old_contact is not None and old_contact.suppressed is True
    new_contact = await db_session.scalar(
        select(Contact).where(
            func.lower(Contact.email) == "marketing001@witofly.com"
        )
    )
    assert new_contact is not None and new_contact.suppressed is False
    assert new_contact.customer_id == customer_id

    await process_inbound(db_session, email_row.id)

    outbox = await _queued_product_list(db_session)
    assert outbox is not None
    assert "marketing001@witofly.com" in outbox.raw_message
    assert "YAC-A110" in outbox.raw_message
