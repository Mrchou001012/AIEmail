from datetime import UTC, date, datetime, timedelta
from decimal import Decimal
from pathlib import Path

import pytest
from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api import (
    InventoryConfirmationRequest,
    InventoryItemRequest,
    confirm_current_inventory,
)
from app.commercial import get_or_create_current_cycle
from app.db import (
    AuditEvent,
    CaseStatus,
    CommercialDataCycle,
    Contact,
    Customer,
    DeliveryStatus,
    EmailMessage,
    InventorySnapshot,
    Job,
    JobStatus,
    Outbox,
    PricePolicy,
    Product,
    Quote,
    SalesCase,
)
from app.imports import generate_templates, import_prices
from app.services import (
    JOB_HANDLERS,
    JobDeferred,
    claim_and_run_job,
    ensure_weekly_commercial_refresh,
    notify_commercial_refresh,
    send_one_outbox,
)
from app.settings import Settings


def _settings(**overrides: object) -> Settings:
    values: dict[str, object] = {
        "demo_mode": False,
        "commercial_gate_enabled": True,
        "commercial_scope": "default",
        "business_timezone": "Asia/Kolkata",
        "business_open_hour": 9,
        "commercial_retry_minutes": 15,
        "mail_transport": "file",
        "auto_send_enabled": False,
        "safe_mode": True,
        "public_base_url": "https://aiemail.example.com",
    }
    values.update(overrides)
    return Settings(_env_file=None, **values)


@pytest.mark.integration
@pytest.mark.asyncio
async def test_weekly_price_replace_is_atomic_and_invalidates_inventory(
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    generate_templates(tmp_path)
    path = tmp_path / "price_list_template.xlsx"

    first = await import_prices(path, db_session, apply=True)
    assert first.ok
    assert first.commercial_cycle_id is not None
    cycle = await db_session.get(CommercialDataCycle, first.commercial_cycle_id)
    assert cycle is not None
    assert cycle.price_status == "CONFIRMED"
    assert cycle.inventory_status == "PENDING"

    product = await db_session.scalar(select(Product).where(Product.code == "WIDGET-100"))
    assert product is not None
    db_session.add(
        InventorySnapshot(
            cycle_id=cycle.id,
            product_id=product.id,
            availability="AVAILABLE",
            source_system="manual",
        )
    )
    cycle.inventory_status = "CONFIRMED"
    cycle.inventory_confirmed_at = datetime.now(UTC)
    await db_session.commit()

    workbook = load_workbook(path)
    columns = {cell.value: cell.column for cell in workbook.active[1]}
    workbook.active.cell(2, columns["standard_price"], 105)
    workbook.save(path)

    second = await import_prices(path, db_session, apply=True, replace_active=True)
    assert second.ok
    assert second.commercial_cycle_id == cycle.id
    policies = (
        (
            await db_session.execute(
                select(PricePolicy)
                .where(PricePolicy.product_id == product.id)
                .order_by(PricePolicy.id)
            )
        )
        .scalars()
        .all()
    )
    assert len(policies) == 2
    assert [row.active for row in policies] == [False, True]
    assert Decimal(policies[-1].standard_price) == Decimal("105.0000")
    assert policies[-1].commercial_cycle_id == cycle.id
    await db_session.refresh(cycle)
    assert cycle.price_status == "CONFIRMED"
    assert cycle.inventory_status == "PENDING"
    assert cycle.inventory_confirmed_at is None
    assert (
        await db_session.scalar(
            select(func.count())
            .select_from(InventorySnapshot)
            .where(InventorySnapshot.cycle_id == cycle.id)
        )
        == 0
    )
    price_audit = await db_session.scalar(
        select(AuditEvent)
        .where(AuditEvent.event_type == "commercial.price_replaced")
        .order_by(AuditEvent.id.desc())
    )
    assert price_audit is not None
    assert price_audit.data["source_hash"] == second.source_hash
    assert price_audit.data["deactivated_policies"] == 1


@pytest.mark.integration
@pytest.mark.asyncio
async def test_price_replace_waits_for_an_in_flight_auto_quote(
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    generate_templates(tmp_path)
    path = tmp_path / "price_list_template.xlsx"
    first = await import_prices(path, db_session, apply=True)
    assert first.ok
    product = await db_session.scalar(select(Product).where(Product.code == "WIDGET-100"))
    policy = await db_session.scalar(
        select(PricePolicy).where(
            PricePolicy.product_id == product.id,
            PricePolicy.active.is_(True),
        )
    )
    cycle = await db_session.get(CommercialDataCycle, first.commercial_cycle_id)
    assert product is not None and policy is not None and cycle is not None
    product_id = product.id

    customer = Customer(
        company_name="In-flight Price Customer",
        auto_send_allowed=True,
        consent_basis="integration test",
    )
    db_session.add(customer)
    await db_session.flush()
    contact = Contact(customer_id=customer.id, name="Buyer", email="inflight@example.com")
    db_session.add(contact)
    await db_session.flush()
    case = SalesCase(
        customer_id=customer.id,
        contact_id=contact.id,
        product_id=product.id,
        currency="USD",
        status=CaseStatus.ACTIVE,
    )
    db_session.add(case)
    await db_session.flush()
    quote = Quote(
        case_id=case.id,
        price_policy_id=policy.id,
        commercial_cycle_id=cycle.id,
        round_number=0,
        unit_price=policy.standard_price,
        currency=policy.currency,
        quantity=100,
        incoterm="EXW",
        payment_term="Prepayment",
        valid_until=cycle.week_end,
        pricing_snapshot={},
    )
    db_session.add(quote)
    await db_session.flush()
    db_session.add(
        Outbox(
            case_id=case.id,
            quote_id=quote.id,
            message_kind="AUTO_QUOTE",
            business_key="in-flight-price-test",
            message_id="<in-flight-price-test@example.com>",
            recipient=contact.email,
            raw_message="claimed auto quote",
            status=DeliveryStatus.CLAIMED,
        )
    )
    await db_session.commit()

    workbook = load_workbook(path)
    columns = {cell.value: cell.column for cell in workbook.active[1]}
    workbook.active.cell(2, columns["standard_price"], 105)
    workbook.save(path)
    replacement = await import_prices(path, db_session, apply=True, replace_active=True)

    assert not replacement.ok
    assert "in flight" in replacement.errors[0]["errors"][0]
    active_prices = (
        (
            await db_session.execute(
                select(PricePolicy).where(
                    PricePolicy.product_id == product_id,
                    PricePolicy.active.is_(True),
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(active_prices) == 1
    assert Decimal(active_prices[0].standard_price) != Decimal("105")


@pytest.mark.integration
@pytest.mark.asyncio
async def test_inventory_confirmation_requires_every_priced_product(
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    generate_templates(tmp_path)
    result = await import_prices(
        tmp_path / "price_list_template.xlsx",
        db_session,
        apply=True,
    )
    assert result.ok
    response = await confirm_current_inventory(
        InventoryConfirmationRequest(
            price_source_ref=result.source_hash,
            items=[
                InventoryItemRequest(
                    product_code="WIDGET-100",
                    availability="AVAILABLE",
                    quantity=Decimal("500"),
                    warehouse="Main",
                )
            ],
            source_system="manual",
            source_ref="weekly-check",
        ),
        "integration-test",
        db_session,
        _settings(),
    )

    assert response["automation_ready"] is True
    assert response["missing_inventory_products"] == []
    assert response["products"][0]["inventory"]["availability"] == "AVAILABLE"


@pytest.mark.integration
@pytest.mark.asyncio
async def test_inventory_confirmation_rejects_a_stale_price_version(
    tmp_path: Path,
    db_session: AsyncSession,
) -> None:
    generate_templates(tmp_path)
    result = await import_prices(
        tmp_path / "price_list_template.xlsx",
        db_session,
        apply=True,
    )
    assert result.ok

    with pytest.raises(HTTPException, match="price batch changed") as exc_info:
        await confirm_current_inventory(
            InventoryConfirmationRequest(
                price_source_ref="stale-price-hash",
                items=[
                    InventoryItemRequest(
                        product_code="WIDGET-100",
                        availability="AVAILABLE",
                    )
                ],
            ),
            "integration-test",
            db_session,
            _settings(),
        )

    assert exc_info.value.status_code == 409
    await db_session.rollback()


@pytest.mark.integration
@pytest.mark.asyncio
async def test_weekly_reminder_is_idempotent_for_many_worker_ticks(
    db_session: AsyncSession,
) -> None:
    settings = _settings(dingtalk_transport="log")
    monday = datetime(2026, 7, 20, 4, 0, tzinfo=UTC)  # 09:30 India

    assert await ensure_weekly_commercial_refresh(db_session, settings, at=monday) is True
    assert await ensure_weekly_commercial_refresh(db_session, settings, at=monday) is False

    jobs = (
        (
            await db_session.execute(
                select(Job).where(Job.kind == "notify_commercial_refresh")
            )
        )
        .scalars()
        .all()
    )
    assert len(jobs) == 1
    assert jobs[0].idempotency_key == "weekly-commercial-refresh:default:2026-07-20"


@pytest.mark.integration
@pytest.mark.asyncio
async def test_completed_cycle_suppresses_a_queued_stale_reminder(
    db_session: AsyncSession,
) -> None:
    cycle = CommercialDataCycle(
        scope="default",
        week_start=date(2026, 7, 20),
        week_end=date(2026, 7, 24),
        price_status="CONFIRMED",
        inventory_status="CONFIRMED",
        reminder_status="PENDING",
    )
    db_session.add(cycle)
    await db_session.commit()

    await notify_commercial_refresh(db_session, cycle.id)

    await db_session.refresh(cycle)
    assert cycle.reminder_status == "NOT_REQUIRED"
    assert cycle.reminder_sent_at is None


@pytest.mark.integration
@pytest.mark.asyncio
async def test_commercial_job_defer_does_not_consume_retry_budget(
    db_session: AsyncSession,
) -> None:
    available_at = datetime.now(UTC) + timedelta(hours=1)

    async def wait_for_business_data(_session: AsyncSession, _payload: dict[str, object]) -> None:
        raise JobDeferred("commercial data waiting: PRICE_CONFIRMATION_PENDING", available_at)

    db_session.add(
        Job(
            kind="commercial-wait-test",
            payload={},
            idempotency_key="commercial-wait-test",
        )
    )
    await db_session.commit()
    JOB_HANDLERS["commercial-wait-test"] = wait_for_business_data
    try:
        assert await claim_and_run_job(db_session, "test-worker", _settings()) is True
    finally:
        JOB_HANDLERS.pop("commercial-wait-test", None)

    job = await db_session.scalar(select(Job).where(Job.idempotency_key == "commercial-wait-test"))
    assert job is not None
    assert job.status is JobStatus.PENDING
    assert job.attempts == 0
    assert job.available_at == available_at
    assert job.last_error.startswith("DEFERRED: commercial data waiting")


@pytest.mark.integration
@pytest.mark.asyncio
async def test_weekend_automated_mail_waits_until_monday(
    db_session: AsyncSession,
) -> None:
    saturday = datetime(2026, 7, 18, 4, 30, tzinfo=UTC)  # Saturday 10:00 India
    row = Outbox(
        business_key="weekend-test",
        message_id="<weekend-test@example.com>",
        recipient="buyer@example.com",
        raw_message="not sent",
        message_kind="GENERAL",
        available_at=saturday,
    )
    db_session.add(row)
    await db_session.commit()

    assert await send_one_outbox(db_session, _settings(), at=saturday) is True
    await db_session.refresh(row)
    assert row.status is DeliveryStatus.PENDING
    assert row.attempts == 0
    assert row.available_at == datetime(2026, 7, 20, 3, 30, tzinfo=UTC)
    assert "until Monday" in (row.last_error or "")


@pytest.mark.integration
@pytest.mark.asyncio
async def test_monday_cancels_frozen_old_quote_and_queues_one_reprice(
    db_session: AsyncSession,
) -> None:
    monday = datetime(2026, 7, 20, 4, 0, tzinfo=UTC)
    old_cycle = CommercialDataCycle(
        scope="default",
        week_start=date(2026, 7, 13),
        week_end=date(2026, 7, 17),
        price_status="CONFIRMED",
        inventory_status="CONFIRMED",
        reminder_status="SENT",
    )
    customer = Customer(
        company_name="Weekly Gate Customer",
        auto_send_allowed=True,
        consent_basis="integration test",
    )
    product = Product(
        code="WEEKLY-100",
        name="Weekly Product",
        unit="kg",
        approved_text_key="widget_100",
    )
    db_session.add_all([old_cycle, customer, product])
    await db_session.flush()
    contact = Contact(customer_id=customer.id, name="Buyer", email="buyer@example.com")
    db_session.add(contact)
    await db_session.flush()
    case = SalesCase(
        customer_id=customer.id,
        contact_id=contact.id,
        product_id=product.id,
        currency="INR",
        status=CaseStatus.ACTIVE,
    )
    policy = PricePolicy(
        commercial_cycle_id=old_cycle.id,
        product_id=product.id,
        currency="INR",
        standard_price=Decimal("100"),
        absolute_floor=Decimal("90"),
        valid_from=old_cycle.week_start,
        source_hash="old-week",
        active=True,
    )
    db_session.add_all([case, policy])
    await db_session.flush()
    inbound = EmailMessage(
        case_id=case.id,
        customer_id=customer.id,
        contact_id=contact.id,
        direction="INBOUND",
        message_id="<weekly-inbound@example.com>",
        from_address=contact.email,
        to_addresses=["sales@example.com"],
        subject="Please quote",
        body_text="Please quote 100 kg WEEKLY-100",
        raw_sha256="weekly-inbound",
        received_at=monday,
    )
    db_session.add(inbound)
    await db_session.flush()
    quote = Quote(
        case_id=case.id,
        price_policy_id=policy.id,
        commercial_cycle_id=old_cycle.id,
        round_number=0,
        unit_price=Decimal("100"),
        currency="INR",
        quantity=100,
        incoterm="EXW",
        payment_term="Prepayment",
        valid_until=old_cycle.week_end,
        pricing_snapshot={},
    )
    db_session.add(quote)
    await db_session.flush()
    outbox = Outbox(
        case_id=case.id,
        quote_id=quote.id,
        message_kind="AUTO_QUOTE",
        business_key=f"inbound-reply:{inbound.id}:quote:{old_cycle.id}",
        message_id="<weekly-outbox@example.com>",
        recipient=contact.email,
        raw_message="frozen old price",
        available_at=monday,
    )
    db_session.add(outbox)
    await db_session.commit()

    assert await send_one_outbox(db_session, _settings(), at=monday) is True
    await db_session.refresh(outbox)
    assert outbox.status is DeliveryStatus.CANCELLED
    assert "commercial data gate cancelled" in (outbox.last_error or "")
    current_cycle = await get_or_create_current_cycle(db_session, _settings(), at=monday)
    reprice_jobs = (
        (
            await db_session.execute(
                select(Job).where(
                    Job.idempotency_key
                    == f"commercial-reprice:inbound:{inbound.id}:cycle:{current_cycle.id}"
                )
            )
        )
        .scalars()
        .all()
    )
    assert len(reprice_jobs) == 1
    assert reprice_jobs[0].status is JobStatus.PENDING
